"""月度考核报表统一入口。

普通更新只允许写入目标班次的状态格和考核结果格。规则同步与新建月份
是仅有的结构性操作；两者都会先基于正式报表做最小范围调整。
"""

from __future__ import annotations

import argparse
import copy
import datetime as dt
import os
import shutil
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from build_report import load_items_from_rules_dir
from fill_report import (
    TEAM_MAP,
    build_column_entries,
    find_assessment_marker_row,
    validate_result_file,
    validate_workbook_schedule,
)
from report_common import (
    concise_team_name,
    expected_run_id,
    load_production_calendar,
    report_date_for_shift,
    save_workbook_atomically,
    shifts_for_date,
)


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
PROJECT_SRC = PROJECT_ROOT / "src"
DEFAULT_RESULTS = PROJECT_ROOT / "assessment_reports"
DEFAULT_RULES = PROJECT_SRC / "dcs_performance" / "rules"

# report.py is also executed directly from the 报表 directory.  Make the
# installed package available for the post-save notification hook without
# changing the existing report helper imports.
if str(PROJECT_SRC) not in sys.path:
    sys.path.insert(0, str(PROJECT_SRC))

from dcs_performance.notification import send_package_email


def _default_report_path(today: dt.date | None = None) -> Path:
    """Select the formal report for the current month without using temp variants."""
    current = today or dt.date.today()
    prefix = f"中控{current.year}年{current.month}月份考核"
    repaired = SCRIPT_DIR / f"{prefix}_修复后.xlsx"
    regular = SCRIPT_DIR / f"{prefix}.xlsx"
    if repaired.is_file():
        return repaired
    if regular.is_file():
        return regular
    return repaired


DEFAULT_REPORT = _default_report_path()

CONTROL_FILL = PatternFill("solid", fgColor="DDEBF7")
SHARED_FILL = PatternFill("solid", fgColor="FFF2CC")
FIELD_FILL = PatternFill("solid", fgColor="E2F0D9")
AUDIENCE_FILLS = {
    "control": CONTROL_FILL,
    "shared": SHARED_FILL,
    "field": FIELD_FILL,
}
TEAM_LABELS = {"A": "甲", "B": "乙", "C": "丙"}
VALID_STATUSES = {"", "是", "免考"}


@dataclass
class MappingEntry:
    sequence: int
    rule_id: str
    point_id: str
    row: int
    audience: str
    effective_date: dt.date | None = None
    inactive_date: dt.date | None = None

    @property
    def key(self) -> tuple[str, str]:
        return self.rule_id, self.point_id

    def applies_on(self, day: dt.date) -> bool:
        if self.effective_date is not None and day < self.effective_date:
            return False
        return self.inactive_date is None or day < self.inactive_date


def _date(value: str, option: str) -> dt.date:
    try:
        return dt.date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError(f"{option} 必须为 YYYY-MM-DD") from exc


def _month(value: str) -> tuple[int, int]:
    try:
        parsed = dt.datetime.strptime(value, "%Y-%m")
    except ValueError as exc:
        raise ValueError("--month 必须为 YYYY-MM") from exc
    return parsed.year, parsed.month


def _team_id(value: str) -> str:
    team = TEAM_MAP.get(value.strip())
    if team is None:
        raise ValueError("班组必须为甲、乙、丙（也可输入 A、B、C）")
    return team


def _parse_optional_date(value: object, field: str) -> dt.date | None:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    try:
        return dt.date.fromisoformat(str(value).strip())
    except ValueError as exc:
        raise ValueError(f"映射表 {field} 不是 YYYY-MM-DD 日期: {value!r}") from exc


def _fallback_audience(sequence: int) -> str:
    if 8 <= sequence <= 13:
        return "shared"
    if 24 <= sequence <= 35:
        return "field"
    return "control"


def _load_mapping(ws: Any) -> list[MappingEntry]:
    entries: list[MappingEntry] = []
    seen_keys: set[tuple[str, str]] = set()
    seen_rows: set[int] = set()
    for r in range(2, ws.max_row + 1):
        rule_id = str(ws.cell(r, 2).value or "").strip()
        point_id = str(ws.cell(r, 3).value or "").strip()
        if not rule_id and not point_id:
            continue
        if not rule_id or not point_id:
            raise ValueError(f"映射表第 {r} 行 rule_id/point_id 不完整")
        try:
            sequence = int(ws.cell(r, 1).value)
            main_row = int(ws.cell(r, 4).value)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"映射表第 {r} 行序号或主表行号无效") from exc
        audience = str(ws.cell(r, 5).value or _fallback_audience(sequence)).strip().lower()
        if audience not in AUDIENCE_FILLS:
            raise ValueError(f"映射表第 {r} 行归属无效: {audience!r}")
        entry = MappingEntry(
            sequence=sequence,
            rule_id=rule_id,
            point_id=point_id,
            row=main_row,
            audience=audience,
            effective_date=_parse_optional_date(ws.cell(r, 6).value, "生效日期"),
            inactive_date=_parse_optional_date(ws.cell(r, 7).value, "停用日期"),
        )
        if entry.key in seen_keys:
            raise ValueError(f"映射表存在重复点位: {entry.rule_id}/{entry.point_id}")
        if entry.row in seen_rows:
            raise ValueError(f"映射表存在重复主表行号: {entry.row}")
        seen_keys.add(entry.key)
        seen_rows.add(entry.row)
        entries.append(entry)
    if not entries:
        raise ValueError("隐藏映射表中没有有效考核项目")
    return sorted(entries, key=lambda item: item.row)


def _open_report(path: Path) -> tuple[Any, Any, Any, list[MappingEntry]]:
    if not path.is_file():
        raise FileNotFoundError(f"正式报表不存在: {path}")
    wb = openpyxl.load_workbook(path, data_only=False)
    if "中控" not in wb.sheetnames or "映射" not in wb.sheetnames:
        raise ValueError("工作簿必须同时包含“中控”和隐藏的“映射”工作表")
    ws = wb["中控"]
    map_ws = wb["映射"]
    mapping = _load_mapping(map_ws)
    return wb, ws, map_ws, mapping


def _find_total_rows(ws: Any) -> tuple[int, int]:
    control = field = None
    for row in range(1, ws.max_row + 1):
        label = str(ws.cell(row, 2).value or "").strip()
        if label == "中控合计：":
            control = row
        elif label == "现场合计：":
            field = row
    if control is None or field is None or field != control + 1:
        raise ValueError("未找到相邻的“中控合计：”和“现场合计：”行")
    return control, field


def _validate_update_layout(ws: Any, map_ws: Any, mapping: list[MappingEntry]) -> None:
    if find_assessment_marker_row(ws) != 3:
        raise ValueError("序号0必须位于第3行")
    control_row, field_row = _find_total_rows(ws)
    expected_rows = list(range(4, control_row))
    actual_rows = [item.row for item in mapping]
    if actual_rows != expected_rows:
        raise ValueError(
            f"映射表项目行必须连续位于第4～{control_row - 1}行，实际为 {actual_rows[:3]}…{actual_rows[-3:]}"
        )
    control_header = field_row + 1
    field_header = control_header + 5
    if [ws.cell(row, 1).value for row in (control_header, field_header)] != ["中控", "现场"]:
        raise ValueError(
            f"中控/现场统计表必须位于合计行之后（当前应为第{control_header}和{field_header}行）"
        )
    if map_ws.sheet_state != "hidden":
        raise ValueError("映射工作表必须保持隐藏")
    if str(ws.freeze_panes) != "F4":
        raise ValueError(f"冻结窗口必须为 F4，实际为 {ws.freeze_panes}")


def _column_context(path: Path) -> tuple[Any, Any, Any, list[MappingEntry], dict[tuple[str, str], int], Any, Any]:
    wb, ws, map_ws, mapping = _open_report(path)
    _validate_update_layout(ws, map_ws, mapping)
    config, calendar, _ = load_production_calendar()
    entries = build_column_entries(ws)
    validate_workbook_schedule(entries, calendar, config)
    columns = {(date_text, team): col for date_text, team, col in entries}
    return wb, ws, map_ws, mapping, columns, config, calendar


def _shift_for_assignment(calendar: Any, config: Any, day: dt.date, team_id: str) -> Any | None:
    for shift in shifts_for_date(calendar, config, day):
        if shift.team_id == team_id:
            return shift
    return None


def _previous_completed_shift(calendar: Any, now: dt.datetime | None = None) -> Any:
    current_time = now or dt.datetime.now()
    current = calendar.shift_for_timestamp(current_time)
    return calendar.shift_for_timestamp(current.start_time - dt.timedelta(microseconds=1))


def _result_path(results_root: Path, shift: Any) -> Path:
    run_id = expected_run_id(shift.start_time, shift.end_time, shift.team_id)
    return results_root / run_id / "result.json"


def _backup_file(path: Path, label: str) -> Path:
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    backup_dir = path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    target = backup_dir / f"{path.stem}_{label}_{stamp}{path.suffix}"
    shutil.copy2(path, target)
    return target


def _backup_result_package(result_json: Path) -> Path | None:
    package = result_json.parent
    if not package.is_dir():
        return None
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    backup_root = package.parent / "backups"
    backup_root.mkdir(parents=True, exist_ok=True)
    target = backup_root / f"{package.name}_{stamp}"
    shutil.copytree(package, target)
    return target


def _run_assessment(shift: Any, results_root: Path, overwrite: bool) -> Path:
    result_json = _result_path(results_root, shift)
    if result_json.is_file() and not overwrite:
        print(f"复用已存在且待校验的原始结果: {result_json}")
        return result_json
    env = os.environ.copy()
    old_pythonpath = env.get("PYTHONPATH")
    env["PYTHONPATH"] = str(PROJECT_SRC) if not old_pythonpath else f"{PROJECT_SRC}{os.pathsep}{old_pythonpath}"
    command = [
        sys.executable,
        "-m",
        "dcs_performance.cli",
        "run",
        "--at",
        (shift.start_time + dt.timedelta(seconds=1)).isoformat(),
        "--output",
        str(results_root),
    ]
    if overwrite:
        command.append("--overwrite")
    subprocess.run(command, cwd=PROJECT_ROOT, env=env, check=True)
    if not result_json.is_file():
        raise RuntimeError(f"考核程序成功返回，但未生成结果文件: {result_json}")
    return result_json


def _write_result(
    report_path: Path,
    day: dt.date,
    team_id: str,
    *,
    force: bool,
    results_root: Path,
    send_email: bool = False,
    email_config: Path | None = None,
    email_state: Path | None = None,
) -> str:
    wb, ws, _, mapping, columns, config, calendar = _column_context(report_path)
    shift = _shift_for_assignment(calendar, config, day, team_id)
    if shift is None:
        print(f"{day.isoformat()} {TEAM_LABELS[team_id]}班不上班，未写入任何内容。")
        return "off"
    col = columns.get((day.isoformat(), team_id))
    if col is None:
        raise ValueError(f"报表中没有 {day.isoformat()} {TEAM_LABELS[team_id]}班对应列")
    status = str(ws.cell(3, col).value or "").strip()
    if status not in VALID_STATUSES:
        raise ValueError(f"{get_column_letter(col)}3 状态无效: {status!r}")
    if status == "免考":
        print(f"跳过 {day.isoformat()} {TEAM_LABELS[team_id]}班：状态为“免考”，需先手工取消免考。")
        return "skipped"
    if status == "是" and not force:
        print(f"跳过 {day.isoformat()} {TEAM_LABELS[team_id]}班：已经处理，保护现有人工修改。")
        return "skipped"

    result_json = _result_path(results_root, shift)
    if force:
        print(f"警告：将强制重算 {day.isoformat()} {TEAM_LABELS[team_id]}班，该列人工修改将被覆盖。")
        workbook_backup = _backup_file(report_path, f"force-{day.isoformat()}-{team_id}")
        result_backup = _backup_result_package(result_json)
        print(f"报表备份: {workbook_backup}")
        print(f"原始结果备份: {result_backup if result_backup else '无（此前没有结果包）'}")

    result_json = _run_assessment(shift, results_root, overwrite=force)
    validated = validate_result_file(result_json, calendar)
    if validated["date"] != day.isoformat() or validated["team_id"] != team_id:
        raise ValueError("结果包日期或班组与目标列不一致")
    active_entries = [item for item in mapping if item.applies_on(day)]
    expected = {item.key for item in active_entries}
    actual = {(rule_id, point_id) for rule_id, point_id, _ in validated["points"]}
    if expected != actual:
        missing = sorted(expected - actual)
        extra = sorted(actual - expected)
        raise ValueError(f"结果与报表项目不一致；缺少 {missing[:5]}，多出 {extra[:5]}")

    # This is the only ordinary-update write scope: target column, rows 3 and mapped item rows.
    for item in mapping:
        ws.cell(item.row, col).value = None
    by_key = {(rule_id, point_id): score for rule_id, point_id, score in validated["points"]}
    penalty_count = 0
    for item in active_entries:
        score = float(by_key[item.key])
        if score <= 0:
            continue
        rounded = round(score, 2)
        ws.cell(item.row, col).value = -int(rounded) if rounded.is_integer() else -rounded
        penalty_count += 1
    ws.cell(3, col).value = "是"
    save_workbook_atomically(wb, report_path)
    if send_email:
        try:
            notification = send_package_email(
                result_json.parent,
                config_path=email_config,
                state_path=email_state,
            )
        except Exception as exc:
            # The Result Package and Excel file are already durable at this
            # point.  Keep them and return a distinct failure status so the
            # caller can retry with dcs-performance send-email --resend.
            print(
                f"邮件发送失败（Excel 和 Result Package 已保存，可稍后单独补发）: {exc}",
                file=sys.stderr,
            )
            return "email_failed"
        print(f"邮件通知已处理: {notification.status}，收件人 {', '.join(notification.recipients)}")
    print(f"已更新 {day.isoformat()} {TEAM_LABELS[team_id]}班：写入 {penalty_count} 项扣分，其余保持空白。")
    return "updated"


def update_command(args: argparse.Namespace) -> int:
    report = args.excel.resolve()
    _, _, _, _, _, config, calendar = _column_context(report)
    assignments: list[tuple[dt.date, str]] = []
    if args.last:
        shift = _previous_completed_shift(calendar)
        assignments.append((report_date_for_shift(shift), shift.team_id))
    elif args.date:
        assignments.append((_date(args.date, "--date"), _team_id(args.team)))
    else:
        start = _date(args.date_from, "--from")
        end = _date(args.date_to, "--to")
        if end < start:
            raise ValueError("--to 不能早于 --from")
        day = start
        while day <= end:
            for shift in shifts_for_date(calendar, config, day):
                assignments.append((day, shift.team_id))
            day += dt.timedelta(days=1)

    if args.force and not args.yes:
        answer = input("强制重算会覆盖目标班次的人工修改，且会先备份。继续？[y/N] ").strip().lower()
        if answer not in {"y", "yes"}:
            print("已取消。")
            return 0

    updated = skipped = email_failed = 0
    for day, team_id in assignments:
        result = _write_result(
            report,
            day,
            team_id,
            force=args.force,
            results_root=args.results.resolve(),
            send_email=args.send_email,
            email_config=args.email_config.resolve() if args.email_config else None,
            email_state=args.email_state.resolve() if args.email_state else None,
        )
        if result == "updated":
            updated += 1
        elif result == "email_failed":
            updated += 1
            email_failed += 1
        else:
            skipped += 1
    print(
        f"处理完成：更新 {updated} 个班次，跳过 {skipped} 个班次。"
        + (f"邮件发送失败 {email_failed} 个班次。" if email_failed else "")
    )
    return 1 if email_failed else 0


def exempt_command(args: argparse.Namespace) -> int:
    day = _date(args.date, "--date")
    team_id = _team_id(args.team)
    report = args.excel.resolve()
    wb, ws, _, mapping, columns, config, calendar = _column_context(report)
    if _shift_for_assignment(calendar, config, day, team_id) is None:
        print(f"{day.isoformat()} {TEAM_LABELS[team_id]}班不上班，未写入任何内容。")
        return 0
    col = columns.get((day.isoformat(), team_id))
    if col is None:
        raise ValueError("目标班次不在当前正式报表中")
    backup = _backup_file(report, f"exempt-{day.isoformat()}-{team_id}")
    for item in mapping:
        ws.cell(item.row, col).value = None
    ws.cell(3, col).value = "免考"
    save_workbook_atomically(wb, report)
    print(f"已设置免考：{day.isoformat()} {TEAM_LABELS[team_id]}班；扣分已清空，统计班次数不计入。")
    print(f"免考前报表备份: {backup}")
    return 0


def _item_dates(item: dict[str, Any]) -> tuple[dt.date | None, dt.date | None]:
    return (
        _parse_optional_date(item.get("effective_date"), "生效日期"),
        _parse_optional_date(item.get("inactive_date"), "停用日期"),
    )


def _item_visible_values(item: dict[str, Any]) -> tuple[str, str, str, str]:
    return str(item["name"]), str(item["range"]), str(item["unit"]), str(item["standard"])


def _ensure_total_row_style(ws: Any, control_row: int, field_row: int) -> None:
    """Keep the field total visually parallel to control with a green fill."""
    ws.row_dimensions[field_row].height = ws.row_dimensions[control_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(control_row, col)
        target = ws.cell(field_row, col)
        target._style = copy.copy(source._style)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)
        target.alignment = copy.copy(source.alignment)
        target.fill = copy.copy(FIELD_FILL)


def _copy_row_style(ws: Any, source_row: int, target_row: int, last_col: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, last_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)
        target.alignment = copy.copy(source.alignment)


def _formula_for_rows(col_letter: str, rows: Iterable[int]) -> str:
    ordered = sorted(rows)
    if not ordered:
        return "=0"
    groups: list[tuple[int, int]] = []
    start = previous = ordered[0]
    for row in ordered[1:]:
        if row == previous + 1:
            previous = row
        else:
            groups.append((start, previous))
            start = previous = row
    groups.append((start, previous))
    refs = [f"{col_letter}{a}" if a == b else f"{col_letter}{a}:{col_letter}{b}" for a, b in groups]
    return f"=SUM({','.join(refs)})"


def _write_mapping(map_ws: Any, entries: list[MappingEntry]) -> None:
    map_ws.delete_rows(1, map_ws.max_row)
    headers = ["序号", "rule_id", "point_id", "主表行号", "归属", "生效日期", "停用日期"]
    for col, value in enumerate(headers, 1):
        map_ws.cell(1, col).value = value
    for index, item in enumerate(sorted(entries, key=lambda entry: entry.row), 1):
        item.sequence = index
        values = [index, item.rule_id, item.point_id, item.row, item.audience, item.effective_date, item.inactive_date]
        for col, value in enumerate(values, 1):
            map_ws.cell(index + 1, col).value = value
        for col in (6, 7):
            map_ws.cell(index + 1, col).number_format = "yyyy-mm-dd"
    map_ws.sheet_state = "hidden"


def _apply_ownership(ws: Any, entries: list[MappingEntry]) -> None:
    for item in entries:
        fill = AUDIENCE_FILLS[item.audience]
        for col in range(1, 6):
            ws.cell(item.row, col).fill = copy.copy(fill)


def _rewrite_totals_and_stats(ws: Any, entries: list[MappingEntry], control_row: int, field_row: int) -> None:
    last_col = ws.max_column
    control_rows = [item.row for item in entries if item.audience in {"control", "shared"}]
    field_rows = [item.row for item in entries if item.audience in {"field", "shared"}]
    ws.cell(control_row, 2).value = "中控合计："
    ws.cell(field_row, 2).value = "现场合计："
    for col in range(6, last_col + 1):
        letter = get_column_letter(col)
        ws.cell(control_row, col).value = _formula_for_rows(letter, control_rows)
        ws.cell(field_row, col).value = _formula_for_rows(letter, field_rows)
    _ensure_total_row_style(ws, control_row, field_row)
    last_letter = get_column_letter(last_col)
    control_header = field_row + 1
    field_header = control_header + 5
    for label, header, total in (("中控", control_header, control_row), ("现场", field_header, field_row)):
        for col, value in enumerate((label, "班次数", "总分", "平均分"), 1):
            ws.cell(header, col).value = value
        for offset, team in enumerate(("甲", "乙", "丙"), 1):
            row = header + offset
            ws.cell(row, 1).value = team
            ws.cell(row, 2).value = f'=COUNTIFS(F$2:{last_letter}$2,A{row},F$3:{last_letter}$3,"是")'
            ws.cell(row, 3).value = f"=SUMIF(F$2:{last_letter}$2,A{row},F${total}:{last_letter}${total})"
            ws.cell(row, 4).value = f'=IF(B{row}=0,"",ROUND(C{row}/B{row},2))'
    ws.print_area = f"A1:{last_letter}{field_header + 3}"


def _current_items() -> list[dict[str, Any]]:
    return load_items_from_rules_dir(DEFAULT_RULES)


def sync_rules_command(args: argparse.Namespace) -> int:
    report = args.excel.resolve()
    wb, ws, map_ws, entries = _open_report(report)
    items = _current_items()
    by_key = {entry.key: entry for entry in entries}
    item_by_key = {(item["rule_id"], item["point_id"]): item for item in items}
    new_items = [item for key, item in item_by_key.items() if key not in by_key]
    warnings = [
        f"{item['rule_id']}/{item['point_id']} 未显式配置归属，已按中控处理"
        for item in new_items
        if item.get("audience_defaulted") and item["audience"] == "control"
    ]
    stopped = [entry for key, entry in by_key.items() if key not in item_by_key and entry.inactive_date is None]
    modified: list[tuple[MappingEntry, dict[str, Any], list[str]]] = []
    for key in sorted(set(by_key) & set(item_by_key)):
        entry, item = by_key[key], item_by_key[key]
        current = tuple(str(ws.cell(entry.row, col).value or "") for col in range(2, 6))
        desired = _item_visible_values(item)
        fields = [name for name, old, new in zip(("位号", "范围", "单位", "扣分标准"), current, desired) if old != new]
        desired_effective, desired_inactive = _item_dates(item)
        if entry.audience != item["audience"]:
            fields.append(f"归属 {entry.audience}->{item['audience']}")
        if entry.effective_date != desired_effective:
            fields.append(f"生效日期 {entry.effective_date or '空白'}->{desired_effective or '空白'}")
        if entry.inactive_date != desired_inactive:
            fields.append(f"停用日期 {entry.inactive_date or '空白'}->{desired_inactive or '空白'}")
        if fields:
            modified.append((entry, item, fields))

    print("规则变更预览：")
    for item in new_items:
        print(f"  新增 {item['rule_id']}/{item['point_id']}（{item['audience']}）")
    for entry, _, fields in modified:
        print(f"  修改 {entry.rule_id}/{entry.point_id}: {', '.join(fields)}")
    for entry in stopped:
        print(f"  停用 {entry.rule_id}/{entry.point_id}")
    for warning in warnings:
        print(f"  提示 {warning}")
    if not new_items and not modified and not stopped:
        print("  无变更。")
        return 0
    if not args.yes:
        answer = input("确认应用以上规则变更？[y/N] ").strip().lower()
        if answer not in {"y", "yes"}:
            print("已取消。")
            return 0

    backup = _backup_file(report, "sync-rules")
    today = dt.date.today()
    for entry, item, _ in modified:
        for col, value in enumerate(_item_visible_values(item), 2):
            ws.cell(entry.row, col).value = value
        entry.audience = item["audience"]
        entry.effective_date, entry.inactive_date = _item_dates(item)
    for entry in stopped:
        entry.inactive_date = today

    control_row, _ = _find_total_rows(ws)
    if new_items:
        ws.insert_rows(control_row, len(new_items))
        for index, item in enumerate(new_items):
            row = control_row + index
            _copy_row_style(ws, control_row - 1, row, ws.max_column)
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).value = None
            effective, inactive = _item_dates(item)
            entries.append(MappingEntry(0, item["rule_id"], item["point_id"], row, item["audience"], effective or today, inactive))
        control_row += len(new_items)
    for index, entry in enumerate(sorted(entries, key=lambda value: value.row), 1):
        entry.sequence = index
        ws.cell(entry.row, 1).value = index
        item = item_by_key.get(entry.key)
        if item is not None:
            for col, value in enumerate(_item_visible_values(item), 2):
                ws.cell(entry.row, col).value = value
    _apply_ownership(ws, entries)
    _write_mapping(map_ws, entries)
    _rewrite_totals_and_stats(ws, entries, control_row, control_row + 1)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    save_workbook_atomically(wb, report)
    print(f"规则已同步，报表备份: {backup}")
    return 0


def repair_layout_command(args: argparse.Namespace) -> int:
    """Repair only the current workbook's total-row presentation and formulas."""
    report = args.excel.resolve()
    wb, ws, map_ws, entries = _open_report(report)
    control_row, field_row = _find_total_rows(ws)
    backup = _backup_file(report, "repair-layout")
    _apply_ownership(ws, entries)
    _rewrite_totals_and_stats(ws, entries, control_row, field_row)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    save_workbook_atomically(wb, report)
    print(f"已修复合计行样式和动态公式，报表备份: {backup}")
    return 0


def _copy_column(ws: Any, source: int, target: int, last_row: int) -> None:
    source_letter = get_column_letter(source)
    target_letter = get_column_letter(target)
    ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width
    ws.column_dimensions[target_letter].hidden = ws.column_dimensions[source_letter].hidden
    for row in range(1, last_row + 1):
        src = ws.cell(row, source)
        dst = ws.cell(row, target)
        if src.has_style:
            dst._style = copy.copy(src._style)
        dst.number_format = src.number_format
        dst.protection = copy.copy(src.protection)
        dst.alignment = copy.copy(src.alignment)


def new_month_command(args: argparse.Namespace) -> int:
    year, month = _month(args.month)
    template = args.excel.resolve()
    output = args.output.resolve() if args.output else SCRIPT_DIR / f"中控{year}年{month}月份考核.xlsx"
    if output.exists():
        raise FileExistsError(f"目标文件已存在，不会覆盖: {output}")
    shutil.copy2(template, output)
    try:
        wb, ws, map_ws, old_entries = _open_report(output)
        items = _current_items()
        previous_keys = {entry.key for entry in old_entries}
        for item in items:
            key = (item["rule_id"], item["point_id"])
            if key not in previous_keys and item.get("audience_defaulted") and item["audience"] == "control":
                print(f"提示: {item['rule_id']}/{item['point_id']} 未显式配置归属，已按中控处理")
        control_row, _ = _find_total_rows(ws)
        current_count = control_row - 4
        desired_count = len(items)
        if desired_count > current_count:
            ws.insert_rows(control_row, desired_count - current_count)
        elif desired_count < current_count:
            ws.delete_rows(4 + desired_count, current_count - desired_count)
        control_row = 4 + desired_count

        old_last_col = ws.max_column
        target_days = __import__("calendar").monthrange(year, month)[1]
        target_last_col = 5 + target_days * 2
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row == 1 and merged.max_row == 1 and merged.min_col >= 6:
                ws.unmerge_cells(str(merged))
        if target_last_col > old_last_col:
            ws.insert_cols(old_last_col + 1, target_last_col - old_last_col)
            for col in range(old_last_col + 1, target_last_col + 1):
                source = old_last_col - ((target_last_col - col) % 2)
                _copy_column(ws, source, col, ws.max_row)
        elif target_last_col < old_last_col:
            ws.delete_cols(target_last_col + 1, old_last_col - target_last_col)

        config, calendar, _ = load_production_calendar()
        ws.cell(1, 1).value = f"聚酯中控考核积分表（{year}年{month}月）"
        for day_number in range(1, target_days + 1):
            day = dt.date(year, month, day_number)
            start_col = 6 + (day_number - 1) * 2
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 1)
            ws.cell(1, start_col).value = day
            ws.cell(1, start_col).number_format = "m/d;@"
            for offset, shift in enumerate(shifts_for_date(calendar, config, day)):
                ws.cell(2, start_col + offset).value = concise_team_name(config, shift.team_id)
            ws.cell(3, start_col).value = None
            ws.cell(3, start_col + 1).value = None

        previous = {entry.key: entry for entry in old_entries}
        new_entries: list[MappingEntry] = []
        month_start = dt.date(year, month, 1)
        for index, item in enumerate(items, 1):
            row = 3 + index
            if row > 4 and row > 3 + len(old_entries):
                _copy_row_style(ws, row - 1, row, target_last_col)
            ws.cell(row, 1).value = index
            for col, value in enumerate(_item_visible_values(item), 2):
                ws.cell(row, col).value = value
            for col in range(6, target_last_col + 1):
                ws.cell(row, col).value = None
            old = previous.get((item["rule_id"], item["point_id"]))
            effective, inactive = _item_dates(item)
            new_entries.append(MappingEntry(index, item["rule_id"], item["point_id"], row, item["audience"], effective or (old.effective_date if old else month_start), inactive))

        _apply_ownership(ws, new_entries)
        _write_mapping(map_ws, new_entries)
        _rewrite_totals_and_stats(ws, new_entries, control_row, control_row + 1)
        ws.freeze_panes = "F4"
        map_ws.sheet_state = "hidden"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        save_workbook_atomically(wb, output)
    except Exception:
        output.unlink(missing_ok=True)
        raise
    print(f"已从正式报表复制并创建新月份: {output}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="月度考核报表统一更新工具")
    parser.add_argument("--excel", type=Path, default=DEFAULT_REPORT, help=f"正式报表，默认 {DEFAULT_REPORT.name}")
    subparsers = parser.add_subparsers(dest="command", required=True)

    update = subparsers.add_parser("update", help="生成并写入考核结果")
    target = update.add_mutually_exclusive_group(required=True)
    target.add_argument("--last", action="store_true", help="更新刚结束的上一个班")
    target.add_argument("--date", help="报表日期 YYYY-MM-DD（需同时提供 --team）")
    target.add_argument("--from", dest="date_from", help="范围起始日期 YYYY-MM-DD（需同时提供 --to）")
    update.add_argument("--to", dest="date_to", help="范围结束日期 YYYY-MM-DD")
    update.add_argument("--team", help="甲、乙、丙")
    update.add_argument("--force", action="store_true", help="备份后强制重算，免考仍不覆盖")
    update.add_argument("--yes", action="store_true", help="跳过强制重算确认提示")
    update.add_argument("--results", type=Path, default=DEFAULT_RESULTS, help="原始考核结果目录")
    update.add_argument(
        "--send-email",
        action="store_true",
        help="Excel 原子保存成功后，从本次 Result Package 发送班次通知",
    )
    update.add_argument("--email-config", type=Path, default=None, help="邮件配置 JSON")
    update.add_argument("--email-state", type=Path, default=None, help="邮件发送状态 JSON")
    update.set_defaults(func=update_command)

    exempt = subparsers.add_parser("exempt", help="设置某班次免考")
    exempt.add_argument("--date", required=True)
    exempt.add_argument("--team", required=True)
    exempt.set_defaults(func=exempt_command)

    sync = subparsers.add_parser("sync-rules", help="预览并同步当前规则")
    sync.add_argument("--yes", action="store_true", help="确认应用预览的规则变更")
    sync.set_defaults(func=sync_rules_command)

    repair = subparsers.add_parser("repair-layout", help="修复现有报表合计行样式和公式")
    repair.set_defaults(func=repair_layout_command)

    new_month = subparsers.add_parser("new-month", help="复制正式报表创建新月份")
    new_month.add_argument("--month", required=True)
    new_month.add_argument("--output", type=Path, default=None)
    new_month.set_defaults(func=new_month_command)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    if args.command == "update":
        if args.date and not args.team:
            parser.error("update --date 必须同时提供 --team")
        if args.date_from and not args.date_to:
            parser.error("update --from 必须同时提供 --to")
        if args.team and not args.date:
            parser.error("--team 只能与 --date 一起使用")
    try:
        return int(args.func(args) or 0)
    except (FileNotFoundError, FileExistsError, OSError, RuntimeError, ValueError, subprocess.CalledProcessError) as exc:
        print(f"错误: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
