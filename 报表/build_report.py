"""build_report.py - 生成中控月度考核 Excel 空表及隐藏映射表.

根据项目现有规则配置（src/dcs_performance/rules/*/config.json），生成指定月份的中控考核积分表：
- 顶部结构：第 1 行合并日期，第 2 行按项目排班每天按“夜班、白班”生成两个成绩格；
- 考核状态：序号 0 行记录每个班次是否已经导入有效 Result Package；
- 考核项：严格读取所有规则目录中的当前启用点位；
- 底部统计：按实际考核项数量和数据列数量生成合计、班组天数、总分、均分公式；
- 隐藏 Sheet【映射】：记录序号、rule_id、point_id 与主表行号的绑定关系，平时隐藏。
"""

from __future__ import annotations

import argparse
import calendar
import datetime
import json
import sys
from collections.abc import Mapping
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from report_common import (
    concise_team_name,
    load_production_calendar,
    save_workbook_atomically,
    shifts_for_date,
)

# ---------------------------------------------------------------------------
# 报表表现层样式与尺寸定义（符合生产管理专业视觉规范）
# ---------------------------------------------------------------------------
FONT_FAMILY = "微软雅黑"

# 字体
FONT_TITLE = Font(name=FONT_FAMILY, size=14, bold=True, color="FFFFFF")
FONT_HEADER_LEFT = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
FONT_HEADER_DATE = Font(name=FONT_FAMILY, size=10, bold=True, color="1F4E78")
FONT_HEADER_SHIFT = Font(name=FONT_FAMILY, size=10, bold=True, color="1F4E78")
FONT_BODY = Font(name=FONT_FAMILY, size=10, color="000000")
FONT_BOLD = Font(name=FONT_FAMILY, size=10, bold=True, color="000000")
FONT_PENALTY = Font(name=FONT_FAMILY, size=10, bold=True, color="9C0006")

# 填充
FILL_TITLE = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
FILL_HEADER_LEFT = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
FILL_HEADER_DATE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_HEADER_SHIFT = PatternFill(start_color="EDF2F8", end_color="EDF2F8", fill_type="solid")
FILL_SUMMARY = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_STAT_HEADER = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
FILL_STAT_FIELD_HEADER = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
FILL_PENALTY = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_CONTROL_POINT = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
FILL_SHARED_POINT = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_FIELD_POINT = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

# 对齐
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_STANDARD = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 边框边线
SIDE_THIN_GRAY = Side(border_style="thin", color="D9D9D9")
SIDE_THIN_LIGHT = Side(border_style="thin", color="E0E0E0")
SIDE_NAVY_MEDIUM = Side(border_style="medium", color="2F5597")
SIDE_NAVY_THIN = Side(border_style="thin", color="2F5597")
SIDE_NAVY_DOUBLE = Side(border_style="double", color="2F5597")
SIDE_HEADER_DIVIDER = Side(border_style="thin", color="5B9BD5")
SIDE_DATE_BORDER = Side(border_style="thin", color="8EA9DB")
SIDE_DAY_SEP = Side(border_style="medium", color="B4C6E7")

# 行高与列宽尺寸
HEIGHT_TITLE = 28.0
HEIGHT_HEADER = 24.0
HEIGHT_DATA = 22.0
HEIGHT_ASSESSMENT = 22.0
HEIGHT_SUMMARY = 23.0
HEIGHT_SPACER = 16.0
HEIGHT_STAT_HEADER = 24.0
HEIGHT_STAT_ROW = 22.0

WIDTH_SEQ = 6.0
WIDTH_NAME = 21.0
WIDTH_RANGE = 18.0
WIDTH_UNIT = 8.0
WIDTH_STANDARD = 30.0
WIDTH_SHIFT = 5.5


# 报表点位只从项目规则配置读取，不在脚本内维护副本。

# 现场专属点与中控/现场共用点使用稳定的业务身份，不依赖 Excel 行号。
SHARED_REPORT_POINTS = {
    ("analog_limit_exceedance", "TIC-012022"),
    ("analog_limit_exceedance", "TIC-015009"),
    ("analog_limit_exceedance", "TIC-117117"),
    ("analog_limit_exceedance", "TIC-217117"),
    ("analog_limit_exceedance", "TIC-117001"),
    ("analog_limit_exceedance", "TIC-217001"),
}

FIELD_REPORT_POINTS = {
    ("persistent_high_alarm", "LA-115077"),
    ("persistent_high_alarm", "LA-115177"),
    ("persistent_high_alarm", "LA-117075"),
    ("persistent_high_alarm", "LA-215077"),
    ("persistent_high_alarm", "LA-215177"),
    ("persistent_high_alarm", "LA-217075"),
    ("pump_flow_compliance", "117P01"),
    ("pump_flow_compliance", "115P05"),
    ("pump_flow_compliance", "115P03"),
    ("pump_flow_compliance", "217P01"),
    ("pump_flow_compliance", "215P05"),
    ("pump_flow_compliance", "215P03"),
}


def report_audience(rule_id: str, point_id: str) -> str:
    key = (rule_id, point_id)
    if key in SHARED_REPORT_POINTS:
        return "shared"
    if key in FIELD_REPORT_POINTS:
        return "field"
    return "control"


def _sum_formula(column: str, rows: list[int]) -> str:
    """Build a readable SUM formula from a sorted list of included rows."""
    if not rows:
        return "=0"
    groups: list[tuple[int, int]] = []
    start = previous = rows[0]
    for row in rows[1:]:
        if row == previous + 1:
            previous = row
            continue
        groups.append((start, previous))
        start = previous = row
    groups.append((start, previous))
    references = [
        f"{column}{start}" if start == end else f"{column}{start}:{column}{end}"
        for start, end in groups
    ]
    return f"=SUM({','.join(references)})"


def _mapping(value: object, context: str) -> Mapping[str, Any]:
    if not isinstance(value, Mapping):
        raise ValueError(f"{context} 必须是对象")
    return value


def _number(value: object, context: str) -> int | float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{context} 必须是数字")
    return value


def _format_number(value: object, context: str) -> str:
    number = _number(value, context)
    if isinstance(number, float) and number.is_integer():
        return str(int(number))
    return str(number)


def _duration_text(value: object, context: str, *, short: bool = False) -> str:
    seconds = _number(value, context)
    if seconds < 0 or int(seconds) != seconds:
        raise ValueError(f"{context} 必须是非负整数秒")
    seconds = int(seconds)
    suffix_minute = "分" if short else "分钟"
    if seconds % 3600 == 0:
        return f"{seconds // 3600}小时"
    if seconds % 60 == 0:
        return f"{seconds // 60}{suffix_minute}"
    return f"{seconds}秒"


def _score(scores: Mapping[str, Any], point_id: str, key: str, default: object) -> int | float:
    point_scores = scores.get(point_id, {})
    if isinstance(point_scores, Mapping):
        value = point_scores.get(key, default)
    else:
        value = point_scores
    return _number(value, f"scoring.{point_id}.{key}")


def _direction_standard(
    directions: list[tuple[str, object, object]],
    *,
    prefix: str,
    short_duration: bool = False,
) -> str:
    if not directions:
        return ""
    if len(directions) == 2 and directions[0][1:] == directions[1][1:]:
        return (
            f"{prefix}>{_duration_text(directions[0][1], 'duration', short=short_duration)}"
            f"扣{_format_number(directions[0][2], 'score')}分/次"
        )
    labels = {"low": "低限", "high": "高限"}
    return "；".join(
        f"{labels.get(direction, direction)}>{_duration_text(duration, 'duration', short=short_duration)}"
        f"扣{_format_number(score, 'score')}分/次"
        for direction, duration, score in directions
    )


def _build_report_item(
    rule_id: str,
    config: Mapping[str, Any],
    point: Mapping[str, Any],
) -> dict[str, Any] | None:
    """Convert one enabled project point into the concise visible row model."""
    point_id = point["id"]
    scoring = _mapping(config.get("scoring", {}), f"{rule_id}.scoring")
    default_score = scoring.get("default_score_per_event", 1)

    if rule_id == "analog_limit_exceedance":
        point_scores = _mapping(
            scoring.get("by_point_event_type", {}),
            f"{rule_id}.scoring.by_point_event_type",
        )
        directions: list[tuple[str, object, object]] = []
        limits: dict[str, object] = {}
        for direction, score_key in (("low", "low_limit"), ("high", "high_limit")):
            raw = point.get(direction)
            if raw is None:
                continue
            direction_config = _mapping(raw, f"{rule_id}.{point_id}.{direction}")
            if direction_config.get("enabled", True) is False:
                continue
            if "limit" not in direction_config:
                raise ValueError(f"{rule_id}.{point_id}.{direction}.limit 缺失")
            limits[direction] = direction_config["limit"]
            directions.append(
                (
                    direction,
                    direction_config.get("min_duration_seconds", 0),
                    _score(point_scores, point_id, score_key, default_score),
                )
            )
        if not directions:
            return None
        if "low" in limits and "high" in limits:
            range_text = (
                f"{_format_number(limits['low'], f'{point_id}.low.limit')}"
                f"～{_format_number(limits['high'], f'{point_id}.high.limit')}"
            )
        elif "low" in limits:
            range_text = f"≥{_format_number(limits['low'], f'{point_id}.low.limit')}"
        else:
            range_text = f"≤{_format_number(limits['high'], f'{point_id}.high.limit')}"
        unit = point.get("unit")
        if not isinstance(unit, str) or not unit:
            unit = "℃" if point_id.startswith("T") else ("dl/g" if point_id.startswith("V") else "%")
        return {
            "rule_id": rule_id,
            "point_id": point_id,
            "name": point_id,
            "range": range_text,
            "unit": unit,
            "standard": _direction_standard(directions, prefix="超限"),
        }

    if rule_id == "component_viscosity_control":
        assessment = _mapping(point.get("assessment", {}), f"{rule_id}.{point_id}.assessment")
        if "low_limit" not in assessment or "high_limit" not in assessment:
            raise ValueError(f"{rule_id}.{point_id}.assessment 上下限缺失")
        point_scores = _mapping(scoring.get("by_point", {}), f"{rule_id}.scoring.by_point")
        directions = [
            (
                "low",
                assessment.get("min_duration_seconds", 0),
                _score(point_scores, point_id, "viscosity_low", default_score),
            ),
            (
                "high",
                assessment.get("min_duration_seconds", 0),
                _score(point_scores, point_id, "viscosity_high", default_score),
            ),
        ]
        return {
            "rule_id": rule_id,
            "point_id": point_id,
            "name": point_id,
            "range": (
                f"{_format_number(assessment['low_limit'], f'{point_id}.low_limit')}"
                f"～{_format_number(assessment['high_limit'], f'{point_id}.high_limit')}"
            ),
            "unit": str(point.get("unit", "MPa")),
            "standard": _direction_standard(directions, prefix="超限"),
        }

    if rule_id == "flow_balance_compliance":
        point_scores = _mapping(scoring.get("by_point", {}), f"{rule_id}.scoring.by_point")
        directions = [
            ("low", point.get("min_duration_seconds", 0), _score(point_scores, point_id, "flow_low", default_score)),
            ("high", point.get("min_duration_seconds", 0), _score(point_scores, point_id, "flow_high", default_score)),
        ]
        return {
            "rule_id": rule_id,
            "point_id": point_id,
            "name": point_id,
            "range": (
                f"{_format_number(point.get('low_limit'), f'{point_id}.low_limit')}"
                f"～{_format_number(point.get('high_limit'), f'{point_id}.high_limit')}"
            ),
            "unit": str(point.get("unit", "偏差")),
            "standard": _direction_standard(directions, prefix="偏差超限"),
        }

    if rule_id == "level_rate_compliance":
        point_scores = _mapping(scoring.get("by_point", {}), f"{rule_id}.scoring.by_point")
        directions = [
            ("low", point.get("persistence_seconds", 0), _score(point_scores, point_id, "rate_down", default_score)),
            ("high", point.get("persistence_seconds", 0), _score(point_scores, point_id, "rate_up", default_score)),
        ]
        return {
            "rule_id": rule_id,
            "point_id": point_id,
            "name": point_id,
            "range": (
                f"{_format_number(point.get('lower_rate'), f'{point_id}.lower_rate')}"
                f"～{_format_number(point.get('upper_rate'), f'{point_id}.upper_rate')}"
            ),
            "unit": str(point.get("unit", "%/h")),
            "standard": _direction_standard(directions, prefix="速率超限"),
        }

    if rule_id == "persistent_high_alarm":
        point_scores = _mapping(scoring.get("by_point", {}), f"{rule_id}.scoring.by_point")
        point_score = point_scores.get(point_id, default_score)
        if isinstance(point_score, Mapping):
            point_score = point_score.get("high_alarm", point_score.get("score", default_score))
        return {
            "rule_id": rule_id,
            "point_id": point_id,
            "name": point_id,
            "range": "0（高报=1）",
            "unit": str(point.get("unit", "状态")),
            "standard": (
                f"高报>{_duration_text(config.get('parameters', {}).get('threshold_seconds', 0), 'threshold_seconds')}"
                f"扣{_format_number(point_score, f'{point_id}.score')}分/次"
            ),
        }

    if rule_id == "pump_flow_compliance":
        event_scores = _mapping(scoring.get("by_event_type", {}), f"{rule_id}.scoring.by_event_type")
        low_score = _number(event_scores.get("low_flow", default_score), f"{rule_id}.low_flow")
        switch_score = _number(event_scores.get("switch_timeout", default_score), f"{rule_id}.switch_timeout")
        return {
            "rule_id": rule_id,
            "point_id": point_id,
            "name": point_id,
            "range": (
                f"正常≥{_format_number(point.get('normal_min_flow'), f'{point_id}.normal_min_flow')}, "
                f"切泵≥{_format_number(point.get('switching_min_flow'), f'{point_id}.switching_min_flow')}"
            ),
            "unit": str(point.get("unit", "t/h")),
            "standard": (
                f"低流量扣{_format_number(low_score, f'{point_id}.low_flow')}分，"
                f"切泵超时>{_duration_text(point.get('max_switch_duration_seconds'), f'{point_id}.max_switch_duration_seconds', short=True)}"
                f"扣{_format_number(switch_score, f'{point_id}.switch_timeout')}分"
            ),
        }

    return {
        "rule_id": rule_id,
        "point_id": point_id,
        "name": point_id,
        "range": str(point.get("range", "")),
        "unit": str(point.get("unit", "")),
        "standard": str(point.get("standard", f"每次扣{_format_number(default_score, f'{rule_id}.default_score')}分")),
    }


def load_items_from_rules_dir(rules_dir: Path | None) -> list[dict[str, Any]]:
    """Read the enabled report points directly from every project rule config.

    A report must fail when the source inventory is unavailable or malformed;
    silently falling back to a copied list is how stale report rows survived.
    """
    if rules_dir is None or not rules_dir.is_dir():
        raise FileNotFoundError(f"规则配置目录不存在: {rules_dir}")

    items: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    candidates = sorted(
        path
        for path in rules_dir.iterdir()
        if path.is_dir() and not path.name.startswith("_")
    )
    if not candidates:
        raise ValueError(f"规则配置目录为空: {rules_dir}")

    for rule_dir in candidates:
        config_path = rule_dir / "config.json"
        rule_path = rule_dir / "rule.py"
        if not config_path.is_file() or not rule_path.is_file():
            raise ValueError(f"规则目录缺少 rule.py 或 config.json: {rule_dir}")
        try:
            config = _mapping(
                json.loads(config_path.read_text(encoding="utf-8")),
                str(config_path),
            )
        except json.JSONDecodeError as exc:
            raise ValueError(f"规则配置不是有效 JSON: {config_path}: {exc}") from exc
        rule_id = config.get("id")
        if not isinstance(rule_id, str) or not rule_id:
            raise ValueError(f"规则配置缺少 id: {config_path}")
        if rule_id != rule_dir.name:
            raise ValueError(f"规则目录名与 id 不一致: {rule_dir.name} != {rule_id}")
        enabled = config.get("enabled", True)
        if not isinstance(enabled, bool):
            raise ValueError(f"{rule_id}.enabled 必须是布尔值")
        if not enabled:
            continue

        parameters = _mapping(config.get("parameters", {}), f"{rule_id}.parameters")
        raw_points = parameters.get("points", [])
        if not isinstance(raw_points, list):
            raise ValueError(f"{rule_id}.parameters.points 必须是数组")
        for index, raw_point in enumerate(raw_points):
            point = _mapping(raw_point, f"{rule_id}.parameters.points[{index}]")
            point_id = point.get("id")
            if not isinstance(point_id, str) or not point_id:
                raise ValueError(f"{rule_id}.parameters.points[{index}].id 缺失")
            point_enabled = point.get("enabled", True)
            if not isinstance(point_enabled, bool):
                raise ValueError(f"{rule_id}.{point_id}.enabled 必须是布尔值")
            if not point_enabled:
                continue
            key = (rule_id, point_id)
            if key in seen:
                raise ValueError(f"规则点重复: {rule_id}/{point_id}")
            seen.add(key)
            item = _build_report_item(rule_id, config, point)
            if item is not None:
                configured_audience = point.get("audience", config.get("audience"))
                if configured_audience is None:
                    item["audience"] = report_audience(rule_id, point_id)
                    item["audience_defaulted"] = True
                else:
                    if configured_audience not in {"control", "field", "shared"}:
                        raise ValueError(
                            f"{rule_id}.{point_id}.audience 必须为 control、field 或 shared"
                        )
                    item["audience"] = configured_audience
                    item["audience_defaulted"] = False
                item["effective_date"] = point.get(
                    "effective_date", config.get("effective_date")
                )
                item["inactive_date"] = point.get(
                    "inactive_date", config.get("inactive_date")
                )
                items.append(item)

    if not items:
        raise ValueError("当前项目配置没有启用的报表点位")
    return items


def apply_report_styles(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    days_in_month: int | None = None,
    item_count: int | None = None,
    team_count: int | None = None,
    item_audiences: list[str] | None = None,
) -> None:
    """Apply layout, ownership colors, split totals, and print settings."""
    last_col_idx = ws.max_column
    if days_in_month is None:
        days_in_month = (last_col_idx - 5) // 2
    if days_in_month <= 0:
        raise ValueError(f"无法根据表格列数解析有效天数: max_column={last_col_idx}")

    total_data_cols = days_in_month * 2
    last_col_idx = 5 + total_data_cols
    last_col_letter = get_column_letter(last_col_idx)

    assessment_row = 3 if (
        ws.cell(3, 1).value == 0
        and str(ws.cell(3, 2).value or "").strip() in {"当班是否考核", "当天是否考核"}
    ) else None
    data_start_row = 4 if assessment_row is not None else 3

    first_total_row: int | None = None
    if item_count is None:
        for r in range(3, ws.max_row + 1):
            val = str(ws.cell(r, 2).value or "").strip()
            if "合计" in val:
                first_total_row = r
                break
        if first_total_row is None:
            raise ValueError("无法在表格中找到合计行")
        item_count = first_total_row - data_start_row
    else:
        first_total_row = data_start_row + item_count

    data_end_row = data_start_row + item_count - 1
    control_total_row = first_total_row
    field_total_row = control_total_row + 1

    if team_count is None:
        team_count = 3

    control_stat_header_row = field_total_row + 2
    field_stat_header_row = control_stat_header_row + team_count + 2
    last_stat_row = field_stat_header_row + team_count

    if item_audiences is None:
        item_audiences = []
        for row in range(data_start_row, data_end_row + 1):
            sequence = ws.cell(row, 1).value
            if isinstance(sequence, int) and 8 <= sequence <= 13:
                item_audiences.append("shared")
            elif isinstance(sequence, int) and 24 <= sequence <= 35:
                item_audiences.append("field")
            else:
                item_audiences.append("control")
    if len(item_audiences) != item_count:
        raise ValueError("考核点归属数量与报表行数不一致")

    # 1. 行高设置
    ws.row_dimensions[1].height = HEIGHT_TITLE
    ws.row_dimensions[2].height = HEIGHT_HEADER
    if assessment_row is not None:
        ws.row_dimensions[assessment_row].height = HEIGHT_ASSESSMENT
    for r in range(data_start_row, data_end_row + 1):
        ws.row_dimensions[r].height = HEIGHT_DATA
    ws.row_dimensions[control_total_row].height = HEIGHT_SUMMARY
    ws.row_dimensions[field_total_row].height = HEIGHT_SUMMARY
    ws.row_dimensions[field_total_row + 1].height = HEIGHT_SPACER
    if team_count > 0:
        for header_row in (control_stat_header_row, field_stat_header_row):
            ws.row_dimensions[header_row].height = HEIGHT_STAT_HEADER
            for idx in range(team_count):
                ws.row_dimensions[header_row + 1 + idx].height = HEIGHT_STAT_ROW
        ws.row_dimensions[field_stat_header_row - 1].height = HEIGHT_SPACER

    # 2. 列宽设置
    ws.column_dimensions["A"].width = WIDTH_SEQ
    ws.column_dimensions["B"].width = WIDTH_NAME
    ws.column_dimensions["C"].width = WIDTH_RANGE
    ws.column_dimensions["D"].width = WIDTH_UNIT
    ws.column_dimensions["E"].width = WIDTH_STANDARD
    for col in range(6, last_col_idx + 1):
        ws.column_dimensions[get_column_letter(col)].width = WIDTH_SHIFT

    # 3. 标题行样式 (Row 1)
    ws.cell(1, 1).font = FONT_TITLE
    ws.cell(1, 1).alignment = ALIGN_CENTER
    for c in range(1, 6):
        cell = ws.cell(1, c)
        cell.fill = FILL_TITLE
        cell.border = Border(
            top=Side(border_style="thin", color="1F4E78"),
            bottom=Side(border_style="thin", color="1F4E78"),
            left=Side(border_style="thin", color="1F4E78") if c == 1 else None,
            right=SIDE_NAVY_MEDIUM if c == 5 else None,
        )

    for day in range(1, days_in_month + 1):
        start_col = 6 + (day - 1) * 2
        end_col = start_col + 1
        for c in range(start_col, end_col + 1):
            cell = ws.cell(1, c)
            cell.fill = FILL_HEADER_DATE
            cell.border = Border(
                top=SIDE_DATE_BORDER,
                bottom=SIDE_DATE_BORDER,
                left=SIDE_DAY_SEP if day > 1 and c == start_col else (SIDE_NAVY_MEDIUM if c == start_col else None),
                right=SIDE_DAY_SEP if c == end_col else None,
            )
        date_cell = ws.cell(1, start_col)
        date_cell.font = FONT_HEADER_DATE
        date_cell.alignment = ALIGN_CENTER
        date_cell.number_format = "m/d;@"

    # 4. 表头行样式 (Row 2)
    for c in range(1, 6):
        cell = ws.cell(2, c)
        cell.font = FONT_HEADER_LEFT
        cell.fill = FILL_HEADER_LEFT
        cell.alignment = ALIGN_CENTER
        cell.border = Border(
            top=Side(border_style="thin", color="2F5597"),
            bottom=SIDE_NAVY_MEDIUM,
            left=Side(border_style="thin", color="2F5597") if c == 1 else SIDE_HEADER_DIVIDER,
            right=SIDE_NAVY_MEDIUM if c == 5 else SIDE_HEADER_DIVIDER,
        )

    for day in range(1, days_in_month + 1):
        start_col = 6 + (day - 1) * 2
        end_col = start_col + 1
        for idx, col in enumerate((start_col, end_col)):
            cell = ws.cell(2, col)
            cell.font = FONT_HEADER_SHIFT
            cell.fill = FILL_HEADER_SHIFT
            cell.alignment = ALIGN_CENTER
            cell.border = Border(
                top=SIDE_DATE_BORDER,
                bottom=SIDE_NAVY_MEDIUM,
                left=SIDE_NAVY_MEDIUM if col == 6 else (SIDE_DAY_SEP if idx == 0 else SIDE_THIN_LIGHT),
                right=SIDE_DAY_SEP if idx == 1 else SIDE_THIN_LIGHT,
            )

    # 5. 当班考核状态行
    if assessment_row is not None:
        for c in range(1, 6):
            cell = ws.cell(assessment_row, c)
            cell.font = FONT_BOLD if c in (1, 2) else FONT_BODY
            cell.fill = FILL_HEADER_SHIFT
            cell.alignment = ALIGN_CENTER
            cell.border = Border(
                top=SIDE_THIN_GRAY,
                bottom=SIDE_NAVY_THIN,
                left=SIDE_THIN_GRAY,
                right=SIDE_NAVY_MEDIUM if c == 5 else SIDE_THIN_GRAY,
            )
        for day in range(1, days_in_month + 1):
            start_col = 6 + (day - 1) * 2
            end_col = start_col + 1
            for idx, col in enumerate((start_col, end_col)):
                cell = ws.cell(assessment_row, col)
                cell.font = FONT_BOLD
                cell.fill = FILL_HEADER_SHIFT
                cell.alignment = ALIGN_CENTER
                cell.border = Border(
                    top=SIDE_THIN_LIGHT,
                    bottom=SIDE_NAVY_THIN,
                    left=SIDE_NAVY_MEDIUM if col == 6 else (SIDE_DAY_SEP if idx == 0 else SIDE_THIN_LIGHT),
                    right=SIDE_DAY_SEP if idx == 1 else SIDE_THIN_LIGHT,
                )

    # 6. 点位信息与成绩区域
    ownership_fills = {
        "control": FILL_CONTROL_POINT,
        "shared": FILL_SHARED_POINT,
        "field": FILL_FIELD_POINT,
    }
    for row in range(data_start_row, data_end_row + 1):
        audience = item_audiences[row - data_start_row]
        for c in range(1, 6):
            cell = ws.cell(row, c)
            cell.font = FONT_BODY
            cell.fill = ownership_fills[audience]
            cell.alignment = ALIGN_STANDARD if c == 5 else ALIGN_CENTER
            cell.border = Border(
                top=SIDE_THIN_GRAY,
                bottom=SIDE_THIN_GRAY,
                left=SIDE_THIN_GRAY,
                right=SIDE_NAVY_MEDIUM if c == 5 else SIDE_THIN_GRAY,
            )

        for day in range(1, days_in_month + 1):
            start_col = 6 + (day - 1) * 2
            end_col = start_col + 1
            for idx, col in enumerate((start_col, end_col)):
                cell = ws.cell(row, col)
                cell.alignment = ALIGN_CENTER
                cell.border = Border(
                    top=SIDE_THIN_LIGHT,
                    bottom=SIDE_THIN_LIGHT,
                    left=SIDE_NAVY_MEDIUM if col == 6 else (SIDE_DAY_SEP if idx == 0 else SIDE_THIN_LIGHT),
                    right=SIDE_DAY_SEP if idx == 1 else SIDE_THIN_LIGHT,
                )
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.font = FONT_PENALTY
                    cell.fill = FILL_PENALTY
                elif cell.value is not None:
                    cell.font = FONT_BODY

    # 7. 中控/现场班次合计行
    for total_row, fill in (
        (control_total_row, FILL_CONTROL_POINT),
        (field_total_row, FILL_FIELD_POINT),
    ):
        for c in range(1, last_col_idx + 1):
            cell = ws.cell(total_row, c)
            cell.font = FONT_BOLD
            cell.fill = fill
            cell.alignment = ALIGN_CENTER
            cell.border = Border(
                top=SIDE_NAVY_THIN,
                bottom=SIDE_NAVY_DOUBLE if total_row == field_total_row else SIDE_NAVY_THIN,
                left=SIDE_NAVY_MEDIUM if c == 6 else SIDE_THIN_GRAY,
                right=SIDE_NAVY_MEDIUM if c == 5 else SIDE_THIN_GRAY,
            )

    # 8. 中控/现场月度统计区域
    if team_count > 0:
        for header_row, header_fill in (
            (control_stat_header_row, FILL_STAT_HEADER),
            (field_stat_header_row, FILL_STAT_FIELD_HEADER),
        ):
            for c in range(1, 5):
                cell = ws.cell(header_row, c)
                cell.font = FONT_HEADER_LEFT
                cell.fill = header_fill
                cell.alignment = ALIGN_CENTER
                cell.border = Border(
                    top=SIDE_NAVY_MEDIUM,
                    bottom=SIDE_NAVY_MEDIUM,
                    left=SIDE_NAVY_MEDIUM if c == 1 else SIDE_HEADER_DIVIDER,
                    right=SIDE_NAVY_MEDIUM if c == 4 else SIDE_HEADER_DIVIDER,
                )

            for idx in range(team_count):
                stat_row = header_row + 1 + idx
                is_last = idx == team_count - 1
                for c in range(1, 5):
                    cell = ws.cell(stat_row, c)
                    cell.alignment = ALIGN_CENTER
                    cell.border = Border(
                        top=SIDE_THIN_GRAY,
                        bottom=SIDE_NAVY_MEDIUM if is_last else SIDE_THIN_GRAY,
                        left=SIDE_NAVY_MEDIUM if c == 1 else SIDE_THIN_GRAY,
                        right=SIDE_NAVY_MEDIUM if c == 4 else SIDE_THIN_GRAY,
                    )
                    cell.font = FONT_BOLD if c == 1 else FONT_BODY
                    if c == 2:
                        cell.number_format = "0"
                    elif c in (3, 4):
                        cell.number_format = "0.00"

    # 9. 视图与冻结窗格
    ws.views.sheetView[0].showGridLines = False
    ws.freeze_panes = "F4" if assessment_row is not None else "F3"

    # 10. 页面与打印设置
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.print_title_rows = "1:3" if assessment_row is not None else "1:2"
    ws.print_title_cols = "A:E"
    ws.print_area = f"A1:{last_col_letter}{last_stat_row}"
    ws.page_margins = PageMargins(
        left=0.5,
        right=0.5,
        top=0.6,
        bottom=0.6,
        header=0.3,
        footer=0.3,
    )

    # 11. 条件格式：实际成绩区域小于 0 的成绩浅红底深红字体
    ws.conditional_formatting = openpyxl.formatting.formatting.ConditionalFormattingList()
    score_range = f"F{data_start_row}:{last_col_letter}{data_end_row}"
    penalty_rule = CellIsRule(
        operator="lessThan",
        formula=["0"],
        stopIfTrue=True,
        fill=FILL_PENALTY,
        font=FONT_PENALTY,
    )
    ws.conditional_formatting.add(score_range, penalty_rule)


def beautify_workbook(excel_path: Path, sheet_name: str = "中控") -> None:
    """对已存在的月度考核 Excel 文件进行表现层样式美化，保留考核数据与公式."""
    if not excel_path.is_file():
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")
    wb = openpyxl.load_workbook(str(excel_path), data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表不存在: {sheet_name}")
    ws = wb[sheet_name]
    apply_report_styles(ws)
    save_workbook_atomically(wb, excel_path)
    print(f"成功美化报表样式: {excel_path}")


def build_monthly_assessment_workbook(
    year: int,
    month: int,
    output_file: Path,
    rules_dir: Path | None = None,
    schedule_path: Path | None = None,
    overwrite: bool = False,
) -> None:
    """创建指定年月的考核 Excel 工作簿."""
    if output_file.exists():
        if output_file.is_dir():
            raise ValueError(f"输出路径不是文件: {output_file}")
        if not overwrite:
            raise FileExistsError(f"输出文件已存在，如需替换请加 --overwrite: {output_file}")
    items = load_items_from_rules_dir(rules_dir)
    schedule_config, production_calendar, _ = load_production_calendar(schedule_path)
    item_count = len(items)
    days_in_month = calendar.monthrange(year, month)[1]
    team_ids = list(schedule_config.team_names)
    if not team_ids:
        raise ValueError("项目排班配置没有班组")

    wb = openpyxl.Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    ws = wb.active
    ws.title = "中控"

    # 1. 写入 A1:E1 大标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.cell(1, 1, f"聚酯中控考核积分表（{year}年{month}月）")

    # 2. 写入 A2:E2 表头
    headers_left = ["序号", "位号", "考核范围", "单位", "考核分值与标准"]
    for c, h in enumerate(headers_left, 1):
        ws.cell(2, c, h)

    # 3. 每天按“夜班、白班”写入项目排班实际生成的两个班次（从第 6 列 F 开始）
    for day in range(1, days_in_month + 1):
        current_date = datetime.date(year, month, day)
        day_shifts = shifts_for_date(production_calendar, schedule_config, current_date)
        start_col = 6 + (day - 1) * 2
        end_col = start_col + len(day_shifts) - 1

        # 第 1 行合并居中日期
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        date_cell = ws.cell(1, start_col, current_date)
        date_cell.number_format = "m/d;@"

        # 第 2 行写当天实际在岗的班组，公共排班函数已按夜班、白班排序
        for idx, shift in enumerate(day_shifts):
            col = start_col + idx
            ws.cell(2, col, concise_team_name(schedule_config, shift.team_id))

    total_data_cols = days_in_month * 2
    last_col_idx = 5 + total_data_cols
    last_col_letter = get_column_letter(last_col_idx)

    # 4. 序号 0 行记录每个班次是否已经完成考核结果导入
    assessment_row = 3
    ws.cell(assessment_row, 1, 0)
    ws.cell(assessment_row, 2, "当班是否考核")

    # 5. 写入当前配置中的启用考核项目
    data_start_row = 4
    data_end_row = data_start_row + item_count - 1

    for idx, it in enumerate(items, 1):
        row = data_start_row + idx - 1
        ws.cell(row, 1, idx)
        ws.cell(row, 2, it["name"])
        ws.cell(row, 3, it["range"])
        ws.cell(row, 4, it["unit"])
        ws.cell(row, 5, it["standard"])
        for col in range(6, last_col_idx + 1):
            ws.cell(row, col, None)

    # 6. 中控/现场分别合计。共用点同时进入两边。
    control_total_row = data_end_row + 1
    field_total_row = control_total_row + 1
    ws.cell(control_total_row, 2, "中控合计：")
    ws.cell(field_total_row, 2, "现场合计：")
    control_rows = [
        data_start_row + idx
        for idx, item in enumerate(items)
        if item["audience"] in {"control", "shared"}
    ]
    field_rows = [
        data_start_row + idx
        for idx, item in enumerate(items)
        if item["audience"] in {"field", "shared"}
    ]
    for col in range(6, last_col_idx + 1):
        col_letter = get_column_letter(col)
        ws.cell(control_total_row, col, _sum_formula(col_letter, control_rows))
        ws.cell(field_total_row, col, _sum_formula(col_letter, field_rows))

    # 7. 中控/现场分别统计班组月度成绩
    control_stat_header_row = field_total_row + 2
    field_stat_header_row = control_stat_header_row + len(team_ids) + 2
    for label, stat_header_row, total_row in (
        ("中控", control_stat_header_row, control_total_row),
        ("现场", field_stat_header_row, field_total_row),
    ):
        for col, value in enumerate((label, "天数", "总分", "均分"), 1):
            ws.cell(stat_header_row, col, value)
        for idx, team_id in enumerate(team_ids):
            stat_row = stat_header_row + 1 + idx
            ws.cell(stat_row, 1, concise_team_name(schedule_config, team_id))
            ws.cell(
                stat_row,
                2,
                f'=COUNTIFS(F$2:{last_col_letter}$2,A{stat_row},F${assessment_row}:{last_col_letter}${assessment_row},"是")',
            )
            ws.cell(
                stat_row,
                3,
                f"=SUMIF(F$2:{last_col_letter}$2,A{stat_row},F${total_row}:{last_col_letter}${total_row})",
            )
            ws.cell(stat_row, 4, f"=IF(B{stat_row}=0,0,ROUND(C{stat_row}/B{stat_row},2))")

    # 应用全套表现层样式系统
    apply_report_styles(
        ws,
        days_in_month=days_in_month,
        item_count=item_count,
        team_count=len(team_ids),
        item_audiences=[item["audience"] for item in items],
    )

    # 8. 创建隐藏 Sheet【映射】
    ws_map = wb.create_sheet(title="映射")
    ws_map.sheet_state = "hidden"

    font_map_header = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    fill_map_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    map_headers = ["序号", "rule_id", "point_id", "主表行号", "归属", "生效日期", "停用日期"]
    for c, h in enumerate(map_headers, 1):
        cell = ws_map.cell(1, c, h)
        cell.font = font_map_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill_map_header

    for idx, it in enumerate(items, 1):
        row = 1 + idx
        target_row = data_start_row + idx - 1
        ws_map.cell(row, 1, idx).alignment = Alignment(horizontal="center", vertical="center")
        ws_map.cell(row, 2, it["rule_id"]).alignment = Alignment(horizontal="left", vertical="center")
        ws_map.cell(row, 3, it["point_id"]).alignment = Alignment(horizontal="left", vertical="center")
        ws_map.cell(row, 4, target_row).alignment = Alignment(horizontal="center", vertical="center")
        ws_map.cell(row, 5, it["audience"]).alignment = Alignment(horizontal="center", vertical="center")
        ws_map.cell(row, 6, it.get("effective_date")).alignment = Alignment(horizontal="center", vertical="center")
        ws_map.cell(row, 7, it.get("inactive_date")).alignment = Alignment(horizontal="center", vertical="center")

    ws_map.column_dimensions["A"].width = 8
    ws_map.column_dimensions["B"].width = 30
    ws_map.column_dimensions["C"].width = 25
    ws_map.column_dimensions["D"].width = 12
    ws_map.column_dimensions["E"].width = 12
    ws_map.column_dimensions["F"].width = 14
    ws_map.column_dimensions["G"].width = 14

    save_workbook_atomically(wb, output_file)
    print(f"成功创建月度考核表: {output_file}（{item_count} 个考核项，{days_in_month} 天）")


def main() -> None:
    print(
        "错误: build_report.py 已停用直接重建入口。"
        "请使用 python 报表\\report.py new-month 或 sync-rules。",
        file=sys.stderr,
    )
    raise SystemExit(2)

    # 保留下方旧参数解析代码，供历史开发环境查阅；不会再从命令行执行。
    parser = argparse.ArgumentParser(description="生成中控月度考核 Excel 空表及映射配置")
    parser.add_argument(
        "--month",
        default=None,
        help="目标月份，格式 YYYY-MM，如 2026-09",
    )
    parser.add_argument(
        "--output",
        "-o",
        default=None,
        help="输出 Excel 文件路径，默认中控{YYYY}年{M}月份考核.xlsx",
    )
    parser.add_argument(
        "--rules-dir",
        default=None,
        help="规则配置目录（可选，默认从项目 src/dcs_performance/rules 读取）",
    )
    parser.add_argument(
        "--schedule",
        default=None,
        help="项目排班配置（可选，默认读取 src/dcs_performance/shifts/performance_schedule.json）",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="允许替换已存在的输出文件",
    )
    parser.add_argument(
        "--beautify",
        default=None,
        help="对已存在的 Excel 报表进行表现层样式美化（不改动考核数据与公式）",
    )

    args = parser.parse_args()

    if args.beautify:
        target_path = Path(args.beautify).resolve()
        try:
            beautify_workbook(target_path)
        except (FileNotFoundError, OSError, RuntimeError, ValueError) as exc:
            print(f"错误: {exc}", file=sys.stderr)
            sys.exit(1)
        return

    if not args.month:
        print("错误: 缺少必需参数 --month (例如 --month 2026-09) 或 --beautify (例如 --beautify path/to/file.xlsx)", file=sys.stderr)
        sys.exit(1)

    try:
        dt = datetime.datetime.strptime(args.month, "%Y-%m")
    except ValueError:
        print("错误: --month 格式必须为 YYYY-MM，如 2026-09", file=sys.stderr)
        sys.exit(1)

    if args.output:
        out_path = Path(args.output).resolve()
    else:
        out_path = Path(f"中控{dt.year}年{dt.month}月份考核.xlsx").resolve()

    rules_dir = None
    if args.rules_dir:
        rules_dir = Path(args.rules_dir).resolve()
    else:
        candidate = Path(__file__).resolve().parent.parent / "src" / "dcs_performance" / "rules"
        rules_dir = candidate

    schedule_path = Path(args.schedule).resolve() if args.schedule else None
    try:
        build_monthly_assessment_workbook(
            dt.year,
            dt.month,
            out_path,
            rules_dir=rules_dir,
            schedule_path=schedule_path,
            overwrite=args.overwrite,
        )
    except (FileNotFoundError, OSError, RuntimeError, ValueError) as exc:
        print(f"错误: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
