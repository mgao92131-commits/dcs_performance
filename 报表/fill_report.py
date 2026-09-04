"""fill_report.py - 将考核结果 JSON 填入中控月度考核 Excel 表.

核心逻辑：
1. 从 Excel 隐藏 Sheet【映射】动态读取 (rule_id, point_id) -> 主表行号；
2. 从 Excel 主表【中控】第 1 行与第 2 行动态解析 (日期, 班组) -> 列号；
3. 校验项目生产排班、Result Package 版本、完整点位集合和汇总数据；
4. 成功导入的班次在序号 0 行标记“是”；
5. 将 score 写成负分，score=0 清除旧值，保持单元格原有格式与公式不变；
6. 通过临时文件原子替换保存 Excel 文件。
"""

from __future__ import annotations

import argparse
import datetime
import json
import math
import re
import sys
from collections.abc import Mapping
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from report_common import (
    expected_run_id,
    load_production_calendar,
    parse_local_datetime,
    report_date_for_shift,
    save_workbook_atomically,
    shifts_for_date,
)


TEAM_MAP = {
    "A": "A",
    "B": "B",
    "C": "C",
    "甲": "A",
    "乙": "B",
    "丙": "C",
    "甲班": "A",
    "乙班": "B",
    "丙班": "C",
}


def parse_date_value(val: Any) -> str | None:
    """将单元格值解析为 YYYY-MM-DD 字符串."""
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")

    s = str(val).strip()
    if not s:
        return None

    # 匹配 2026-09-01 或 2026/09/01
    m = re.fullmatch(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        y, mth, d = m.groups()
        try:
            return datetime.date(int(y), int(mth), int(d)).isoformat()
        except ValueError:
            return None

    # 匹配 9/1 或 9-1 或 9月1日
    m2 = re.fullmatch(r"(\d{1,2})[月/-](\d{1,2})日?", s)
    if m2:
        mth, d = m2.groups()
        try:
            datetime.date(2000, int(mth), int(d))
        except ValueError:
            return None
        return f"{int(mth):02d}-{int(d):02d}"

    return None


def build_column_entries(
    ws: Any,
    start_col: int = 6,
    default_year: int = 2026,
) -> list[tuple[str, str, int]]:
    """解析主表中的 (日期, 项目班组, 列号)，不接受无法识别的表头。"""
    entries: list[tuple[str, str, int]] = []
    seen: set[tuple[str, str]] = set()
    current_date_str: str | None = None

    # 先从 A1 标题或其他单元格提取默认年份，防止 m/d 没有年份
    a1_val = str(ws.cell(1, 1).value or "")
    m_year = re.search(r"(\d{4})年", a1_val)
    if m_year:
        default_year = int(m_year.group(1))

    for col in range(start_col, ws.max_column + 1):
        # 第 1 行日期（合并单元格仅左上角有值，需向前传递）
        val1 = ws.cell(1, col).value
        parsed1 = parse_date_value(val1)
        if parsed1 is not None:
            if len(parsed1) == 5:  # MM-DD
                current_date_str = f"{default_year}-{parsed1}"
            else:
                current_date_str = parsed1
                default_year = int(current_date_str.split("-")[0])

        # 第 2 行班组：甲/乙/丙 或 A/B/C
        raw_header = ws.cell(2, col).value
        val2 = str(raw_header or "").strip()
        team_id = TEAM_MAP.get(val2)

        if current_date_str and raw_header not in (None, "") and not team_id:
            raise ValueError(f"第 2 行存在无法识别的班组表头: {val2!r}（第 {col} 列）")
        if current_date_str and team_id:
            key = (current_date_str, team_id)
            if key in seen:
                raise ValueError(f"主表存在重复的日期/班组列: {current_date_str} {team_id}")
            seen.add(key)
            entries.append((current_date_str, team_id, col))

    if not entries:
        return []
    dates = sorted({entry[0] for entry in entries})
    first_date = datetime.date.fromisoformat(dates[0])
    last_date = datetime.date.fromisoformat(dates[-1])
    expected_dates = {
        (first_date + datetime.timedelta(days=offset)).isoformat()
        for offset in range((last_date - first_date).days + 1)
    }
    if set(dates) != expected_dates:
        missing = sorted(expected_dates - set(dates))
        raise ValueError(f"主表日期列不连续，缺少: {', '.join(missing)}")

    return entries


def build_column_map(
    ws: Any,
    start_col: int = 6,
    default_year: int = 2026,
) -> dict[tuple[str, str], int]:
    """根据主表第 1、2 行构建 (日期, 项目班组) -> 列号映射。"""
    column_map: dict[tuple[str, str], int] = {}
    for date_str, team_id, col in build_column_entries(ws, start_col, default_year):
        column_map[(date_str, team_id)] = col

    return column_map


def build_row_map(ws_map: Any) -> dict[tuple[str, str], int]:
    """从隐藏 Sheet【映射】读取 (rule_id, point_id) -> 主表行号映射."""
    row_map: dict[tuple[str, str], int] = {}
    for r in range(2, ws_map.max_row + 1):
        rule_id = str(ws_map.cell(r, 2).value or "").strip()
        point_id = str(ws_map.cell(r, 3).value or "").strip()
        row_num = ws_map.cell(r, 4).value
        if rule_id and point_id and row_num is not None:
            try:
                row_num = int(row_num)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"映射表第 {r} 行主表行号无效: {row_num!r}") from exc
            if row_num < 3:
                raise ValueError(f"映射表第 {r} 行主表行号必须不小于 3: {row_num}")
            key = (rule_id, point_id)
            if key in row_map:
                raise ValueError(f"映射表存在重复点位: {rule_id}/{point_id}")
            row_map[key] = row_num
    return row_map


def find_assessment_marker_row(ws: Any) -> int:
    """Locate the required sequence-zero row used to mark imported shifts."""
    for row in range(3, min(ws.max_row, 20) + 1):
        sequence = ws.cell(row, 1).value
        label = str(ws.cell(row, 2).value or "").strip()
        if sequence == 0 and label in {"当班是否考核", "当天是否考核"}:
            return row
    raise ValueError("主表缺少序号 0 的‘当班是否考核’行")


def collect_result_files(target_path: Path) -> list[Path]:
    """收集待处理的 result.json 文件列表."""
    if target_path.is_file():
        return [target_path]
    if target_path.is_dir():
        return sorted(
            path
            for path in target_path.rglob("result.json")
            if path.is_file()
            and not any(part.startswith((".tmp-", ".backup-")) for part in path.parts)
        )
    return []


def _require_mapping(value: object, context: str) -> Mapping[str, Any]:
    if not isinstance(value, Mapping):
        raise ValueError(f"{context} 必须是对象")
    return value


def _require_text(value: object, context: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{context} 必须是非空字符串")
    return value.strip()


def _require_finite_number(value: object, context: str) -> int | float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{context} 必须是数字")
    if not math.isfinite(float(value)):
        raise ValueError(f"{context} 必须是有限数字")
    return value


def _require_nonnegative_int(value: object, context: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        raise ValueError(f"{context} 必须是非负整数")
    return value


def validate_workbook_schedule(
    entries: list[tuple[str, str, int]],
    production_calendar: Any,
    schedule_config: Any,
) -> set[str]:
    """Ensure the workbook columns are exactly the project's two shifts/day."""
    if not entries:
        raise ValueError("主表中未解析到任何日期与班组列")

    by_date: dict[str, list[str]] = {}
    for date_str, team_id, _ in entries:
        by_date.setdefault(date_str, []).append(team_id)

    dates = sorted(by_date)
    first_date = datetime.date.fromisoformat(dates[0])
    last_date = datetime.date.fromisoformat(dates[-1])
    expected_dates = {
        (first_date + datetime.timedelta(days=offset)).isoformat()
        for offset in range((last_date - first_date).days + 1)
    }
    if set(dates) != expected_dates:
        missing = sorted(expected_dates - set(dates))
        raise ValueError(f"主表日期列不完整，缺少: {', '.join(missing)}")

    for date_str in dates:
        day = datetime.date.fromisoformat(date_str)
        expected_shifts = shifts_for_date(production_calendar, schedule_config, day)
        expected_teams = [shift.team_id for shift in expected_shifts]
        actual_teams = by_date[date_str]
        if actual_teams != expected_teams:
            raise ValueError(
                f"主表 {date_str} 班次列与项目排班不一致: "
                f"应为 {expected_teams}，实际为 {actual_teams}"
            )

    return set(dates)


def validate_result_file(
    json_file: Path,
    production_calendar: Any,
) -> dict[str, Any]:
    """Validate one Result Package and return only fields used for filling."""
    try:
        data = json.loads(json_file.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise ValueError(f"无法读取结果文件 {json_file}: {exc}") from exc
    data = _require_mapping(data, str(json_file))

    if data.get("schema_version") != "1.0":
        raise ValueError(f"{json_file} 的 schema_version 不是 1.0")
    if data.get("time_basis") != "local":
        raise ValueError(f"{json_file} 的 time_basis 必须为 local")

    shift = _require_mapping(data.get("shift"), f"{json_file}.shift")
    team_id = _require_text(shift.get("team_id"), f"{json_file}.shift.team_id")
    shift_type = _require_text(shift.get("shift_type"), f"{json_file}.shift.shift_type")
    if shift_type not in {"day", "night"}:
        raise ValueError(f"{json_file}.shift.shift_type 无效: {shift_type}")
    start_time = parse_local_datetime(shift.get("start"), f"{json_file}.shift.start")
    end_time = parse_local_datetime(shift.get("end"), f"{json_file}.shift.end")
    expected_shift = production_calendar.shift_for_timestamp(start_time)
    if (
        expected_shift.team_id != team_id
        or expected_shift.shift_type != shift_type
        or expected_shift.start_time != start_time
        or expected_shift.end_time != end_time
    ):
        raise ValueError(
            f"{json_file} 的班次与项目排班不一致: "
            f"结果为 {team_id}/{shift_type}/{start_time.isoformat()}~{end_time.isoformat()}，"
            f"应为 {expected_shift.team_id}/{expected_shift.shift_type}/"
            f"{expected_shift.start_time.isoformat()}~{expected_shift.end_time.isoformat()}"
        )

    run = _require_mapping(data.get("run"), f"{json_file}.run")
    run_id = _require_text(run.get("run_id"), f"{json_file}.run.run_id")
    if run_id != expected_run_id(start_time, end_time, team_id):
        raise ValueError(f"{json_file}.run.run_id 与班次不一致: {run_id}")

    rules = data.get("rules")
    if not isinstance(rules, list):
        raise ValueError(f"{json_file}.rules 必须是数组")

    point_values: list[tuple[str, str, int | float]] = []
    point_keys: set[tuple[str, str]] = set()
    rule_ids: set[str] = set()
    total_event_count = 0
    total_score = 0.0

    for rule_index, raw_rule in enumerate(rules):
        rule = _require_mapping(raw_rule, f"{json_file}.rules[{rule_index}]")
        rule_id = _require_text(rule.get("rule_id"), f"{json_file}.rules[{rule_index}].rule_id")
        if rule_id in rule_ids:
            raise ValueError(f"{json_file} 中规则重复: {rule_id}")
        rule_ids.add(rule_id)
        points = rule.get("points")
        if not isinstance(points, list):
            raise ValueError(f"{json_file}.{rule_id}.points 必须是数组")

        rule_event_count = _require_nonnegative_int(
            rule.get("event_count"), f"{json_file}.{rule_id}.event_count"
        )
        rule_score = _require_finite_number(rule.get("score"), f"{json_file}.{rule_id}.score")
        if rule_score < 0:
            raise ValueError(f"{json_file}.{rule_id}.score 不能为负数")
        calculated_rule_events = 0
        calculated_rule_score = 0.0

        for point_index, raw_point in enumerate(points):
            point = _require_mapping(raw_point, f"{json_file}.{rule_id}.points[{point_index}]")
            point_id = _require_text(
                point.get("point_id"),
                f"{json_file}.{rule_id}.points[{point_index}].point_id",
            )
            key = (rule_id, point_id)
            if key in point_keys:
                raise ValueError(f"{json_file} 中点位重复: {rule_id}/{point_id}")
            point_keys.add(key)
            status = _require_text(
                point.get("status"),
                f"{json_file}.{rule_id}.{point_id}.status",
            )
            if status not in {"normal", "violation"}:
                raise ValueError(f"{json_file}.{rule_id}.{point_id}.status 无效: {status}")
            data_status = _require_text(
                point.get("data_status"),
                f"{json_file}.{rule_id}.{point_id}.data_status",
            )
            if data_status not in {"ok", "partial", "no_data"}:
                raise ValueError(
                    f"{json_file}.{rule_id}.{point_id}.data_status 无效: {data_status}"
                )
            event_count = _require_nonnegative_int(
                point.get("event_count"),
                f"{json_file}.{rule_id}.{point_id}.event_count",
            )
            score = _require_finite_number(
                point.get("score"),
                f"{json_file}.{rule_id}.{point_id}.score",
            )
            if score < 0:
                raise ValueError(f"{json_file}.{rule_id}.{point_id}.score 不能为负数")
            events = point.get("events")
            if not isinstance(events, list):
                raise ValueError(f"{json_file}.{rule_id}.{point_id}.events 必须是数组")
            if len(events) != event_count:
                raise ValueError(
                    f"{json_file}.{rule_id}.{point_id}.events 与 event_count 不一致"
                )
            expected_status = "violation" if event_count else "normal"
            if status != expected_status:
                raise ValueError(
                    f"{json_file}.{rule_id}.{point_id}.status 与 event_count 不一致"
                )

            point_values.append((rule_id, point_id, score))
            calculated_rule_events += event_count
            calculated_rule_score += float(score)

        if calculated_rule_events != rule_event_count:
            raise ValueError(f"{json_file}.{rule_id} 的 event_count 汇总不一致")
        if not math.isclose(calculated_rule_score, float(rule_score), abs_tol=1e-9):
            raise ValueError(f"{json_file}.{rule_id} 的 score 汇总不一致")
        total_event_count += rule_event_count
        total_score += float(rule_score)

    summary = _require_mapping(data.get("summary"), f"{json_file}.summary")
    summary_rule_count = _require_nonnegative_int(
        summary.get("rule_count"), f"{json_file}.summary.rule_count"
    )
    summary_point_count = _require_nonnegative_int(
        summary.get("point_count"), f"{json_file}.summary.point_count"
    )
    summary_event_count = _require_nonnegative_int(
        summary.get("event_count"), f"{json_file}.summary.event_count"
    )
    summary_score = _require_finite_number(
        summary.get("total_score"), f"{json_file}.summary.total_score"
    )
    if summary_rule_count != len(rules):
        raise ValueError(f"{json_file}.summary.rule_count 汇总不一致")
    if summary_point_count != len(point_values):
        raise ValueError(f"{json_file}.summary.point_count 汇总不一致")
    if summary_event_count != total_event_count:
        raise ValueError(f"{json_file}.summary.event_count 汇总不一致")
    if not math.isclose(total_score, float(summary_score), abs_tol=1e-9):
        raise ValueError(f"{json_file}.summary.total_score 汇总不一致")

    return {
        "path": json_file,
        "run_id": run_id,
        "date": report_date_for_shift(expected_shift).isoformat(),
        "team_id": team_id,
        "points": point_values,
    }


def fill_assessment_report(
    excel_path: Path,
    results_path: Path,
    sheet_name: str = "中控",
    map_sheet_name: str = "映射",
    dry_run: bool = False,
    schedule_path: Path | None = None,
) -> int:
    """将 assessment 结果填入 Excel."""
    if not excel_path.is_file():
        print(f"错误: Excel 文件不存在: {excel_path}", file=sys.stderr)
        return 1

    try:
        wb = openpyxl.load_workbook(str(excel_path), data_only=False)
    except (OSError, ValueError) as exc:
        print(f"错误: 无法打开 Excel 文件 {excel_path}: {exc}", file=sys.stderr)
        return 1

    if sheet_name not in wb.sheetnames:
        print(f"错误: 工作簿中未找到主表 '{sheet_name}'", file=sys.stderr)
        return 1

    if map_sheet_name not in wb.sheetnames:
        print(f"错误: 工作簿中未找到映射表 '{map_sheet_name}'", file=sys.stderr)
        return 1

    ws = wb[sheet_name]
    ws_map = wb[map_sheet_name]

    try:
        row_map = build_row_map(ws_map)
        assessment_marker_row = find_assessment_marker_row(ws)
        column_entries = build_column_entries(ws)
        schedule_config, production_calendar, _ = load_production_calendar(schedule_path)
        workbook_dates = validate_workbook_schedule(
            column_entries,
            production_calendar,
            schedule_config,
        )
    except (FileNotFoundError, OSError, RuntimeError, ValueError) as exc:
        print(f"错误: {exc}", file=sys.stderr)
        return 1

    column_map = {(date_str, team_id): col for date_str, team_id, col in column_entries}

    if not row_map:
        print(f"错误: 映射表 '{map_sheet_name}' 中未解析到任何有效映射规则", file=sys.stderr)
        return 1

    result_files = collect_result_files(results_path)
    if not result_files:
        print(f"提示: 未在指定路径找到任何 result.json 文件: {results_path}")
        return 0

    print(f"找到 {len(result_files)} 个结果文件，开始校验...")
    try:
        results = [validate_result_file(path, production_calendar) for path in result_files]
    except ValueError as exc:
        print(f"错误: {exc}", file=sys.stderr)
        return 1

    expected_keys = set(row_map)
    seen_assignments: set[tuple[str, str]] = set()
    pending: dict[tuple[int, int], int | float | str | None] = {}
    pending_details: dict[tuple[int, int], tuple[str, str, str, str]] = {}
    penalty_count = 0
    clear_count = 0
    marker_count = 0
    skipped_count = 0

    for result in results:
        result_path = result["path"]
        result_date = result["date"]
        team_id = result["team_id"]
        assignment = (result_date, team_id)
        if assignment in seen_assignments:
            print(f"错误: 同一班次存在多个结果文件: {result_date} {team_id}", file=sys.stderr)
            return 1
        seen_assignments.add(assignment)

        result_keys = {(rule_id, point_id) for rule_id, point_id, _ in result["points"]}
        if result_keys != expected_keys:
            missing = sorted(expected_keys - result_keys)
            extra = sorted(result_keys - expected_keys)
            print(
                f"错误: {result_path} 点位全集与报表映射不一致；"
                f"缺少 {missing[:5]}{'…' if len(missing) > 5 else ''}，"
                f"多出 {extra[:5]}{'…' if len(extra) > 5 else ''}",
                file=sys.stderr,
            )
            return 1

        if result_date not in workbook_dates:
            skipped_count += 1
            print(f"跳过非本月班次: {result_date} {team_id} ({result_path.name})")
            continue

        col = column_map.get(assignment)
        if col is None:
            print(f"错误: 结果班次未找到对应表格列: {result_date} {team_id}", file=sys.stderr)
            return 1

        marker_target = (assessment_marker_row, col)
        if ws.cell(row=assessment_marker_row, column=col).value != "是":
            pending[marker_target] = "是"
            pending_details[marker_target] = (result_date, team_id, "report_status", "当班是否考核")
            marker_count += 1

        for rule_id, point_id, score in result["points"]:
            row = row_map[(rule_id, point_id)]
            numeric_score = float(score)
            if numeric_score == 0:
                value_to_write: int | float | None = None
                clear_count += 1
            else:
                rounded_score = round(numeric_score, 2)
                value_to_write = -int(rounded_score) if rounded_score.is_integer() else -rounded_score
                penalty_count += 1

            target = (row, col)
            if target in pending:
                print(f"错误: 多个结果同时写入同一格: {get_column_letter(col)}{row}", file=sys.stderr)
                return 1
            if ws.cell(row=row, column=col).value != value_to_write:
                pending[target] = value_to_write
                pending_details[target] = (result_date, team_id, rule_id, point_id)

    if dry_run:
        for (row, col), value in pending.items():
            date_str, team_id, rule_id, point_id = pending_details[(row, col)]
            display_value = "空白" if value is None else str(value)
            print(
                f"[预览] {get_column_letter(col)}{row} <- {display_value} "
                f"({date_str} {team_id} {rule_id}/{point_id})"
            )
        print(
            f"\n[预览模式] 将更新 {len(pending)} 格（考核标记 {marker_count}，"
            f"扣分 {penalty_count}，清空 {clear_count}）；"
            f"跳过 {skipped_count} 个非本月班次。"
        )
        return 0

    for (row, col), value in pending.items():
        ws.cell(row=row, column=col).value = value
    if pending:
        save_workbook_atomically(wb, excel_path)
    print(
        f"\n填充完成：更新 {len(pending)} 格（考核标记 {marker_count}，"
        f"扣分 {penalty_count}，清空 {clear_count}）；"
        f"跳过 {skipped_count} 个非本月班次。文件已保存至: {excel_path}"
    )

    return 0


def main() -> None:
    print(
        "错误: fill_report.py 已停用，禁止绕过保护流程直接填表。"
        "请使用 python 报表\\report.py update ...",
        file=sys.stderr,
    )
    sys.exit(2)


if __name__ == "__main__":
    main()
