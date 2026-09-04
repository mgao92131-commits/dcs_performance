from __future__ import annotations

import copy
import json
import shutil
import sys
from datetime import date, datetime
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import openpyxl


REPORT_DIR = Path(__file__).resolve().parents[1] / "报表"
if str(REPORT_DIR) not in sys.path:
    sys.path.insert(0, str(REPORT_DIR))

import report  # noqa: E402


FORMAL = REPORT_DIR / "中控2026年9月份考核_修复后.xlsx"


def _copy_report(tmp_path: Path) -> Path:
    target = tmp_path / "report.xlsx"
    shutil.copy2(FORMAL, target)
    return target


def _fake_result_writer(report_path: Path, results_root: Path, day: date, team_id: str):
    wb, ws, _, mapping = report._open_report(report_path)
    config, calendar, _ = report.load_production_calendar()
    shift = report._shift_for_assignment(calendar, config, day, team_id)
    assert shift is not None
    by_rule: dict[str, list[dict[str, object]]] = {}
    for item in mapping:
        by_rule.setdefault(item.rule_id, []).append(
            {
                "point_id": item.point_id,
                "status": "normal",
                "data_status": "ok",
                "event_count": 0,
                "score": 0,
                "events": [],
            }
        )
    rules = [
        {"rule_id": rule_id, "event_count": 0, "score": 0, "points": points}
        for rule_id, points in sorted(by_rule.items())
    ]
    document = {
        "schema_version": "1.0",
        "time_basis": "local",
        "shift": {
            "team_id": team_id,
            "shift_type": shift.shift_type,
            "start": shift.start_time.isoformat(),
            "end": shift.end_time.isoformat(),
        },
        "run": {"run_id": report.expected_run_id(shift.start_time, shift.end_time, team_id)},
        "rules": rules,
        "summary": {
            "rule_count": len(rules),
            "point_count": len(mapping),
            "event_count": 0,
            "total_score": 0,
        },
    }
    path = report._result_path(results_root, shift)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(document, ensure_ascii=False), encoding="utf-8")
    return path


def test_current_total_row_has_field_style_and_preserves_layout():
    wb = openpyxl.load_workbook(FORMAL, data_only=False)
    ws = wb["中控"]
    assert ws.max_row == 49
    assert ws.max_column == 65
    assert ws.row_dimensions[40].height == ws.row_dimensions[39].height
    assert ws["A40"].fill.fgColor.rgb.endswith("E2F0D9")
    assert ws["F40"].fill.fgColor.rgb.endswith("E2F0D9")
    assert ws["A40"].border.bottom.style == ws["A39"].border.bottom.style == "double"
    assert ws["F40"].border.left.style == ws["F39"].border.left.style == "medium"
    assert ws.freeze_panes == "F4"
    assert wb["映射"].sheet_state == "hidden"


def test_stats_show_average_score_instead_of_rank():
    wb = openpyxl.load_workbook(FORMAL, data_only=False)
    ws = wb["中控"]
    assert ws["D41"].value == "平均分"
    assert ws["D46"].value == "平均分"
    for row in (42, 43, 44, 47, 48, 49):
        assert ws.cell(row, 4).value == f'=IF(B{row}=0,"",ROUND(C{row}/B{row},2))'


def test_new_month_all_day_counts_and_dynamic_update(tmp_path):
    for month, days in (("2026-02", 28), ("2028-02", 29), ("2026-04", 30), ("2026-10", 31)):
        output = tmp_path / f"report-{month}.xlsx"
        assert report.main(
            ["--excel", str(FORMAL), "new-month", "--month", month, "--output", str(output)]
        ) == 0
        wb = openpyxl.load_workbook(output, data_only=False)
        ws = wb["中控"]
        assert ws.max_column == 5 + days * 2
        assert ws.max_row == 49
        assert ws.print_area.endswith(f":${report.get_column_letter(ws.max_column)}$49")
        assert len([m for m in ws.merged_cells.ranges if m.min_row == 1 and m.min_col >= 6]) == days
        assert wb["映射"].sheet_state == "hidden"

        results_root = tmp_path / f"results-{month}"
        day = date.fromisoformat(f"{month}-01")
        config, calendar, _ = report.load_production_calendar()
        shift = report.shifts_for_date(calendar, config, day)[0]
        with patch.object(
            report,
            "_run_assessment",
            side_effect=lambda s, root, overwrite: _fake_result_writer(output, root, day, s.team_id),
        ):
            assert report.main(
                [
                    "--excel",
                    str(output),
                    "update",
                    "--date",
                    day.isoformat(),
                    "--team",
                    report.TEAM_LABELS[shift.team_id],
                    "--results",
                    str(results_root),
                ]
            ) == 0
        updated = openpyxl.load_workbook(output, data_only=False)["中控"]
        column = report.build_column_entries(updated)[0][2]
        assert updated.cell(3, column).value == "是"
        assert updated.cell(39, updated.max_column).value.startswith("=SUM(")
        assert updated.cell(40, updated.max_column).value.startswith("=SUM(")


def test_range_update_and_force_backup(tmp_path):
    output = _copy_report(tmp_path)
    results_root = tmp_path / "results"
    config, calendar, _ = report.load_production_calendar()
    assignments = [
        (day, shift.team_id)
        for day in (date(2026, 9, 3), date(2026, 9, 4))
        for shift in report.shifts_for_date(calendar, config, day)
    ]

    def fake_run(shift, root, overwrite):
        return _fake_result_writer(output, root, report.report_date_for_shift(shift), shift.team_id)

    with patch.object(report, "_run_assessment", side_effect=fake_run):
        assert report.main(
            [
                "--excel",
                str(output),
                "update",
                "--from",
                "2026-09-03",
                "--to",
                "2026-09-04",
                "--results",
                str(results_root),
            ]
        ) == 0
    wb = openpyxl.load_workbook(output, data_only=False)
    ws = wb["中控"]
    columns = report.build_column_entries(ws)
    for day, team_id in assignments:
        col = next(col for d, team, col in columns if d == day.isoformat() and team == team_id)
        assert ws.cell(3, col).value == "是"

    # Mark one processed column with a manual score, then force it. The score is overwritten
    # and a backup is created before the simulated assessment runs.
    target_day, target_team = assignments[0]
    col = next(col for d, team, col in columns if d == target_day.isoformat() and team == target_team)
    ws.cell(7, col).value = -99
    wb.save(output)
    with patch.object(report, "_run_assessment", side_effect=fake_run):
        assert report.main(
            [
                "--excel",
                str(output),
                "update",
                "--date",
                target_day.isoformat(),
                "--team",
                report.TEAM_LABELS[target_team],
                "--force",
                "--yes",
                "--results",
                str(results_root),
            ]
        ) == 0
    final = openpyxl.load_workbook(output, data_only=False)["中控"]
    assert final.cell(3, col).value == "是"
    assert final.cell(7, col).value is None
    assert list((tmp_path / "backups").glob("*.xlsx"))


def test_exempt_creates_backup_and_only_changes_target_column(tmp_path):
    output = _copy_report(tmp_path)
    before = openpyxl.load_workbook(output, data_only=False)
    before_ws = before["中控"]
    target_col = next(
        col for day, team, col in report.build_column_entries(before_ws) if day == "2026-09-03" and team == "A"
    )
    before_values = {(r, c): before_ws.cell(r, c).value for r in range(1, 50) for c in range(1, 66)}
    before_styles = {(r, c): before_ws.cell(r, c)._style for r in range(1, 50) for c in range(1, 66)}
    assert report.main(
        ["--excel", str(output), "exempt", "--date", "2026-09-03", "--team", "甲"]
    ) == 0
    after = openpyxl.load_workbook(output, data_only=False)["中控"]
    changed = {
        (r, c)
        for r in range(1, 50)
        for c in range(1, 66)
        if before_values[(r, c)] != after.cell(r, c).value
    }
    expected_changed = {(3, target_col)} | {
        (r, target_col)
        for r in range(4, 39)
        if before_values[(r, target_col)] is not None
    }
    assert changed == expected_changed
    assert all(
        before_styles[(r, c)] == after.cell(r, c)._style
        for r in range(1, 50)
        for c in range(1, 66)
        if (r, c) != (3, target_col)
    )
    assert list((tmp_path / "backups").glob("*.xlsx"))

    protected_hash = output.stat().st_mtime_ns, output.stat().st_size
    with patch.object(report, "_run_assessment", side_effect=AssertionError("免考不应运行考核")):
        assert report.main(
            ["--excel", str(output), "update", "--date", "2026-09-03", "--team", "甲"]
        ) == 0
    assert (output.stat().st_mtime_ns, output.stat().st_size) == protected_hash


def test_sync_detects_rule_dates_new_and_stopped_points(tmp_path, monkeypatch):
    rules = tmp_path / "rules"
    shutil.copytree(report.DEFAULT_RULES, rules)
    monkeypatch.setattr(report, "DEFAULT_RULES", rules)

    # Date-only changes are included in the preview and persisted to the mapping.
    config_path = rules / "analog_limit_exceedance" / "config.json"
    config = json.loads(config_path.read_text(encoding="utf-8"))
    config["effective_date"] = "2026-09-05"
    config_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    dated_report = _copy_report(tmp_path)
    assert report.main(["--excel", str(dated_report), "sync-rules", "--yes"]) == 0
    mapping = report._open_report(dated_report)[3]
    assert next(item for item in mapping if item.point_id == "LICA-012019").effective_date == date(2026, 9, 5)

    # Add one point: it is appended before totals and the resulting layout is accepted by update.
    add_report = _copy_report(tmp_path)
    add_config = json.loads(config_path.read_text(encoding="utf-8"))
    extra = copy.deepcopy(add_config["parameters"]["points"][0])
    extra["id"] = "LIC-NEW-TEST"
    extra["history_tag"] = "LIC-NEW-TEST/PID1/PV.CV"
    add_config["parameters"]["points"].append(extra)
    config_path.write_text(json.dumps(add_config, ensure_ascii=False, indent=2), encoding="utf-8")
    assert report.main(["--excel", str(add_report), "sync-rules", "--yes"]) == 0
    assert report._column_context(add_report)[0] is not None
    add_wb = openpyxl.load_workbook(add_report, data_only=False)
    assert add_wb["中控"].max_row == 50
    assert add_wb["中控"]["B40"].value == "中控合计："
    assert add_wb["中控"]["B41"].value == "现场合计："
    add_config_obj, add_calendar, _ = report.load_production_calendar()
    add_day = date(2026, 9, 6)  # after the configured effective date
    add_shift = report.shifts_for_date(add_calendar, add_config_obj, add_day)[0]
    add_results = tmp_path / "add-results"
    with patch.object(
        report,
        "_run_assessment",
        side_effect=lambda shift, root, overwrite: _fake_result_writer(
            add_report, root, add_day, shift.team_id
        ),
    ):
        assert report.main(
            [
                "--excel",
                str(add_report),
                "update",
                "--date",
                add_day.isoformat(),
                "--team",
                report.TEAM_LABELS[add_shift.team_id],
                "--results",
                str(add_results),
            ]
        ) == 0
    updated_add = openpyxl.load_workbook(add_report, data_only=False)["中控"]
    add_col = next(col for d, team, col in report.build_column_entries(updated_add) if d == add_day.isoformat() and team == add_shift.team_id)
    assert updated_add.cell(3, add_col).value == "是"

    # Remove a point: it remains in the mapping with a stop date instead of deleting history.
    stop_report = _copy_report(tmp_path)
    stop_config_path = rules / "persistent_high_alarm" / "config.json"
    stop_config = json.loads(stop_config_path.read_text(encoding="utf-8"))
    removed_id = stop_config["parameters"]["points"][-1]["id"]
    stop_config["parameters"]["points"].pop()
    stop_config_path.write_text(json.dumps(stop_config, ensure_ascii=False, indent=2), encoding="utf-8")
    assert report.main(["--excel", str(stop_report), "sync-rules", "--yes"]) == 0
    stopped = next(item for item in report._open_report(stop_report)[3] if item.point_id == removed_id)
    assert stopped.inactive_date is not None
    assert report._column_context(stop_report)[0] is not None


def test_previous_completed_shift_uses_previous_slot_at_boundaries():
    config, calendar, _ = report.load_production_calendar()
    at_boundary = datetime(2026, 9, 4, 8, 0)
    previous = report._previous_completed_shift(calendar, at_boundary)
    assert previous.end_time == at_boundary
    assert report.report_date_for_shift(previous) == date(2026, 9, 4)


def test_update_send_email_runs_only_after_atomic_excel_save(tmp_path, monkeypatch):
    output = _copy_report(tmp_path)
    results_root = tmp_path / "results"
    events: list[str] = []

    def fake_run(shift, root, overwrite):
        return _fake_result_writer(output, root, date(2026, 9, 6), shift.team_id)

    original_save = report.save_workbook_atomically

    def save_and_record(wb, path):
        events.append("excel-save")
        return original_save(wb, path)

    def fake_send(package, **kwargs):
        events.append("email-send")
        assert Path(package).name == "20260905T200000_20260906T080000_A"
        return SimpleNamespace(status="sent", recipients=("a@example.com",))

    monkeypatch.setattr(report, "_run_assessment", fake_run)
    monkeypatch.setattr(report, "save_workbook_atomically", save_and_record)
    monkeypatch.setattr(report, "send_package_email", fake_send)
    assert report.main(
        [
            "--excel",
            str(output),
            "update",
            "--date",
            "2026-09-06",
            "--team",
            "甲",
            "--results",
            str(results_root),
            "--send-email",
            "--email-config",
            str(tmp_path / "notification.json"),
        ]
    ) == 0
    assert events == ["excel-save", "email-send"]
    assert (results_root / "20260905T200000_20260906T080000_A" / "result.json").is_file()


def test_update_email_failure_returns_nonzero_without_rollback(tmp_path, monkeypatch, capsys):
    output = _copy_report(tmp_path)
    results_root = tmp_path / "results"

    def fake_run(shift, root, overwrite):
        return _fake_result_writer(output, root, date(2026, 9, 6), shift.team_id)

    monkeypatch.setattr(report, "_run_assessment", fake_run)

    def broken_send(*args, **kwargs):
        raise RuntimeError("SMTP down")

    monkeypatch.setattr(report, "send_package_email", broken_send)
    assert report.main(
        [
            "--excel",
            str(output),
            "update",
            "--date",
            "2026-09-06",
            "--team",
            "甲",
            "--results",
            str(results_root),
            "--send-email",
        ]
    ) == 1
    captured = capsys.readouterr()
    assert "邮件发送失败" in captured.err
    assert "可稍后单独补发" in captured.err
    assert (results_root / "20260905T200000_20260906T080000_A" / "result.json").is_file()
    ws = openpyxl.load_workbook(output, data_only=False)["中控"]
    column = next(
        col
        for day, team, col in report.build_column_entries(ws)
        if day == "2026-09-06" and team == "A"
    )
    assert ws.cell(3, column).value == "是"
